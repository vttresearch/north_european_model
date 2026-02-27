import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import src.utils as utils
from src.pipeline.bb_excel_context import BBExcelBuildContext


class BuildInputExcel:

    PARAM_GNU = [
        'isActive',
        'capacity',
        'conversionCoeff',
        'useInitialGeneration',
        'initialGeneration',
        'maxRampUp',
        'maxRampDown',
        'rampPenalty',
        'rampUpPenalty',
        'rampDownPenalty',
        'rampUpCost',
        'rampDownCost',
        'upperLimitCapacityRatio',
        'unitSize',
        'invCosts',
        'annuityFactor',
        'invEnergyCost',
        'fomCosts',
        'vomCosts',
        'inertia',
        'unitSizeMVA',
        'availabilityCapacityMargin',
        'startCostCold',
        'startCostWarm',
        'startCostHot',
        'startFuelConsCold',
        'startFuelConsWarm',
        'startFuelConsHot',
        'shutdownCost',
        'delay',
        'cb',
        'cv',
    ]

    PARAM_UNIT = [
        'isActive',
        'isSource',
        'isSink',
        'fixedFlow',
        'availability',
        'unitCount',
        'useInitialOnlineStatus',
        'initialOnlineStatus',
        'startColdAfterXhours',
        'startWarmAfterXhours',
        'rampSpeedToMinLoad',
        'rampSpeedFromMinLoad',
        'minOperationHours',
        'minShutdownHours',
        'eff00',
        'eff01',
        'opFirstCross',
        'op00',
        'op01',
        'useTimeseries',
        'useTimeseriesAvailability',
        'investMIP',
        'maxUnitCount',
        'minUnitCount',
        'becomeAvailable',
        'becomeUnavailable',
    ]

    PARAM_GNN = [
        'transferCap',
        'availability',
        'variableTransCost',
        'transferLoss',
        'rampLimit',
        'diffCoeff',
        'diffLosses',
        'transferCapInvLimit',
        'investMIP',
        'invCost',
        'annuityFactor',
    ]

    PARAM_GN = [
        'isActive',
        'nodeBalance',
        'usePrice',
        'energyStoredPerUnitOfState',
        'selfDischargeLoss',
        'boundStart',
        'boundStartOfSamples',
        'boundStartAndEnd',
        'boundStartToEnd',
        'boundEnd',
        'boundEndOfSamples',
        'boundAll',
        'boundSumOverInterval',
        'capacityMargin',
        'storageValueUseTimeSeries',
        'influx',
    ]

    def __init__(
        self,
        context: BBExcelBuildContext,
        logger=None
        ) -> None:

        self.context = context
        self.input_folder = context.input_folder
        self.output_folder = context.output_folder
        self.scen_tags = context.scen_tags
        self.config = context.config
        self.country_codes = self.config["country_codes"]

        # From InputDataPipeline
        self.source_data = context.source_data
        # Global
        self.df_fueldata =     self.source_data.df_fueldata
        self.df_emissiondata = self.source_data.df_emissiondata
        # Country specific
        self.df_transferdata = self.source_data.df_transferdata
        self.df_unitdata =     self.source_data.df_unitdata
        self.df_storagedata =  self.source_data.df_storagedata
        self.df_demanddata =   self.source_data.df_demanddata
        # Custom
        self.df_userconstraintdata = self.source_data.df_userconstraintdata

        # From TimeseriesPipeline
        self.ts_results = context.ts_results

        # unpack ts_domains and ts_domain_pairs
        self.ts_domains = self.ts_results.ts_domains
        self.ts_domain_pairs = self.ts_results.ts_domain_pairs

        # Extract timeseries data with single dot access
        self.ts_storage_limits = {
            key: value 
            for key, value in self.ts_results.secondary_results.items() 
            if key is not None and key.startswith("ts_storage_limits")
        }
        mingen_vars = {
            key: value 
            for key, value in self.ts_results.secondary_results.items() 
            if key is not None and key.startswith("mingen_nodes")
        }
        self.mingen_nodes = [
            item
            for sublist in mingen_vars.values() or [] 
            if isinstance(sublist, list)
            for item in sublist
        ]
        
        self.logger = logger

        # Define the merged output file
        self.output_file = os.path.join(self.output_folder, 'inputData.xlsx')

        # Initiate a flag for successful code excecution
        self.bb_excel_succesfully_built = False



# ------------------------------------------------------
# Functions create and modify p_gnu_io 
# ------------------------------------------------------

    def create_p_gnu_io(
        self,
        df_unitdata: pd.DataFrame
        ) -> pd.DataFrame:
        """
        Creates a DataFrame representing unit input/output connections with parameters.

        This method processes the merged df_unitdata (which already contains type-level
        defaults from merge_unittypedata_into_unitdata) to build a relationship table
        between grids, nodes and units, with associated parameters.

        Removes zeroes.

        Parameters:
        -----------
        df_unitdata : DataFrame
            Merged unit data. Must include 'generator_id' and 'unit' columns,
            grid_input1/grid_output1/... columns, and node_input1/node_output1/...
            columns (all added by build_unit_grid_and_node_columns).
            Type-level parameter defaults are pre-merged via
            merge_unittypedata_into_unitdata().

        Returns:
        --------
        DataFrame
            A multi-index DataFrame with dimensions (grid, node, unit, input_output)
            and parameter columns (capacity, conversionCoeff, vomCosts, etc.)
        """
        if df_unitdata.empty:
            return pd.DataFrame()

        # dimension and parameter columns
        dimensions = ['grid', 'node', 'unit', 'input_output']
        param_gnu = self.PARAM_GNU

        # List to collect the new rows
        rows = []

        # Process each row in the merged df_unitdata.
        for _, cap_row in df_unitdata.iterrows():

            # Fetch unit name
            unit = cap_row['unit']

            # Identify all defined input/output connections for this unit.
            # grid_* columns were added to df_unitdata by build_unit_grid_and_node_columns.
            put_candidates = ['input1', 'input2', 'input3', 'input4', 'input5',
                              'output1', 'output2', 'output3', 'output4', 'output5']
            available_puts = [put for put in put_candidates if f'grid_{put}' in cap_row.index]

            # Process each available input/output connection
            for put in available_puts:
                # Construct looped column names
                grid_col = f"grid_{put}"
                node_col = f"node_{put}"

                # skip if the columns do not exist in unitdata
                if grid_col not in cap_row:
                    self.logger.log_status(f"Missing grid {grid_col} from unit {unit}, not writing the data. "
                               "Check spelling and files.'", level="warn")
                    continue
                if node_col not in cap_row:
                    self.logger.log_status(f"Missing node {node_col} from unit {unit}, not writing the data. "
                               "Check spelling and files.'", level="warn")
                    continue

                # get values from unitdata
                grid = cap_row.get(grid_col)
                node = cap_row.get(node_col)

                # skip undefined / blank grids
                if pd.isna(grid) or grid in ("", "-"):
                    continue

                # Construct base components needed for every row
                base_row = {
                    "grid": grid,
                    "node": node,
                    "unit": unit,
                    "input_output": "input" if put.startswith("input") else "output",
                }

                # Add all other parameters. _ensure_numeric_dtypes guarantees that
                # all param_gnu columns use the '{param}_{put}' form and are Float64
                # with no NA, so a direct Series lookup is sufficient.
                additional_params = {
                    param: cap_row.get(f'{param.lower()}_{put}', 0)
                    for param in param_gnu
                }

                # Append base and additional parameters
                rows.append({**base_row, **additional_params})

        # Construct p_gnu_io
        final_cols = dimensions + param_gnu
        p_gnu_io = pd.DataFrame(rows, columns=final_cols)
        p_gnu_io = utils.fill_numeric_na(utils.standardize_df_dtypes(p_gnu_io))

        #  Remove empty columns except mandatory 'capacity' column
        p_gnu_io = p_gnu_io.drop(columns=[col for col in p_gnu_io.columns
                                          if utils.is_col_empty(p_gnu_io[col]) and col != 'capacity'])

        # Sort by unit, input_output, node in a case-insensitive manner.
        p_gnu_io.sort_values(by=['unit', 'input_output', 'node'], 
                            key=lambda col: col.str.lower() if col.dtype == 'object' else col,
                            inplace=True)

        # Remove zeroes
        p_gnu_io = p_gnu_io.mask(p_gnu_io == 0)

        # create fake MultiIndex
        p_gnu_io = self.create_fake_MultiIndex(p_gnu_io, dimensions)

        return p_gnu_io


    def fill_capacities(
        self, 
        p_gnu_io: pd.DataFrame, 
        p_unit: pd.DataFrame
        ) -> pd.DataFrame:
        """
        Fills missing capacity values of units with a set of rules. 
        Currently calculates missing input capacity, if 
            * 1 input and 1 output, other one without capacity
            * unit has 1 input without capacity, 2 or more outputs with capacity, and no 'cv' parameter
        """

        # Skip processing if either of the source dataframes are empty
        if p_gnu_io.empty or p_unit.empty:
            return pd.DataFrame()

        # Get a flat version of the source dataframes without the fake multi-index
        p_gnu_io_flat = self.drop_fake_MultiIndex(p_gnu_io).copy()
        p_unit_flat = self.drop_fake_MultiIndex(p_unit)

        # Create efficiency lookup dictionary
        eff_columns = [col for col in p_unit_flat.columns if col.startswith('eff')]
        unit_efficiency = {}
        for _, unit_row in p_unit_flat.iterrows():
            unit_efficiency[unit_row['unit']] = unit_row[eff_columns].max()

        # Helper function to get unit's io rows
        def get_unit_rows(unit, io_type):
            return p_gnu_io_flat[
                (p_gnu_io_flat['unit'] == unit) & 
                (p_gnu_io_flat['input_output'] == io_type)
            ]

        # Process each unit only once
        for unit in p_gnu_io_flat['unit'].unique():
            efficiency = unit_efficiency.get(unit, 0)
            if efficiency <= 0:
                continue

            inputs = get_unit_rows(unit, 'input')
            outputs = get_unit_rows(unit, 'output')

            # Rule 1: 1 input and 1 output, other one without capacity
            if len(inputs) == 1 and len(outputs) == 1:
                input_idx = inputs.index[0]
                output_idx = outputs.index[0]
                input_cap = inputs.iloc[0]['capacity']
                output_cap = outputs.iloc[0]['capacity']

                # If input cap zero and value in output, calculate input from output
                if utils.is_val_empty(input_cap) and not utils.is_val_empty(output_cap):
                    p_gnu_io_flat.at[input_idx, 'capacity'] = output_cap / efficiency

                # If output cap zero and value in input, calculate output from input
                elif not utils.is_val_empty(input_cap) and utils.is_val_empty(output_cap):
                    p_gnu_io_flat.at[output_idx, 'capacity'] = input_cap * efficiency

            # Rule 2: 1 input without capacity, 2 or more outputs with capacity (no 'cv')
            elif len(inputs) == 1 and len(outputs) > 1:
                input_idx = inputs.index[0]
                input_cap = inputs.iloc[0]['capacity']

                if utils.is_val_empty(input_cap):
                    # Check both outputs have capacity
                    output_caps = outputs['capacity']
                    if all(~utils.is_val_empty(cap) for cap in output_caps):
                        # Check for 'cv' parameter if column exists
                        skip = False
                        if 'cv' in outputs.columns:
                            cv_values = outputs['cv']
                            if (cv_values > 0).any():
                                skip = True

                        if not skip:
                            total_output = output_caps.sum()
                            p_gnu_io_flat.at[input_idx, 'capacity'] = total_output / efficiency

        # Recreate the fake multi-index
        p_gnu_io = self.create_fake_MultiIndex(p_gnu_io_flat, ['grid', 'node', 'unit', 'input_output'])
        return p_gnu_io


    def drop_redundant_units(
        self, 
        p_gnu_io: pd.DataFrame, 
        p_unit: pd.DataFrame
        ) -> tuple[pd.DataFrame, pd.DataFrame]:
        """
        Drops units that have no capacity and no investment parameters.

        A unit is dropped when both conditions are true:
            1. All capacity values are zero/NaN for every (grid, node, unit) row in p_gnu_io
            2. Investment parameters are zero/NaN: invCosts in p_gnu_io and maxUnitCount in p_unit

        Parameters
        ----------
        p_gnu_io : DataFrame
            Unit input/output table with fake MultiIndex (row 0 = header).
        p_unit : DataFrame
            Unit parameter table with fake MultiIndex (row 0 = header).

        Returns
        -------
        tuple of (DataFrame, DataFrame)
            Filtered p_gnu_io and p_unit with redundant units removed.
        """
        if p_gnu_io.empty or p_unit.empty:
            return p_gnu_io, p_unit

        gnu_flat = self.drop_fake_MultiIndex(p_gnu_io).copy()
        unit_flat = self.drop_fake_MultiIndex(p_unit).copy()

        units_to_drop = []

        for unit in gnu_flat['unit'].unique():
            unit_rows = gnu_flat[gnu_flat['unit'] == unit]

            # Condition 1: all capacity values are zero or NaN
            has_capacity = False
            if 'capacity' in unit_rows.columns:
                has_capacity = unit_rows['capacity'].notna().any() and (unit_rows['capacity'] != 0).any()
            if has_capacity:
                continue

            # Condition 2: invCosts zero/NaN in p_gnu_io AND maxUnitCount zero/NaN in p_unit
            has_inv_costs = False
            if 'invCosts' in unit_rows.columns:
                has_inv_costs = unit_rows['invCosts'].notna().any() and (unit_rows['invCosts'] != 0).any()
            if has_inv_costs:
                continue

            has_max_unit_count = False
            unit_row = unit_flat[unit_flat['unit'] == unit]
            if not unit_row.empty and 'maxUnitCount' in unit_row.columns:
                val = unit_row['maxUnitCount'].iloc[0]
                has_max_unit_count = pd.notna(val) and val != 0
            if has_max_unit_count:
                continue

            units_to_drop.append(unit)

        if units_to_drop:
            for unit in units_to_drop:
                self.logger.log_status(f"Dropping unit '{unit}': zero capacity and no investment parameters.",
                                 level="skip")
            gnu_flat = gnu_flat[~gnu_flat['unit'].isin(units_to_drop)]
            unit_flat = unit_flat[~unit_flat['unit'].isin(units_to_drop)]

            p_gnu_io = self.create_fake_MultiIndex(gnu_flat, ['grid', 'node', 'unit', 'input_output'])
            p_unit = self.create_fake_MultiIndex(unit_flat, ['unit'])

        return p_gnu_io, p_unit


# ------------------------------------------------------
# Functions create p_gnn, p_gn
# ------------------------------------------------------

    def create_p_gnn(
        self, 
        df_transferdata: pd.DataFrame
        ) -> pd.DataFrame:
        """
        Build p_gnn by looping over 'dimensions' and 'param_gnn'.
        Special cases:
          - transferCap: from export/import capacity depending on direction
          - rampLimit: forward = row['ramplimit'] (default 0),
                       reverse = scaled by (export/import) if both > 0
        """
        if df_transferdata.empty:
            return pd.DataFrame()

        dimensions = ['grid', 'from_node', 'to_node']
        param_gnn = self.PARAM_GNN

        def get_or_default(row, param):
            """Return row[param] (case-insensitive) or 0 if absent."""
            key = param.lower()
            return row.get(key, 0)

        rows = []
        for _, row in df_transferdata.iterrows():
            # domains
            grid      = row.get('grid')
            from_node = row.get('from_node')
            to_node   = row.get('to_node')
            # specific values
            export_cap = get_or_default(row, 'export_capacity')
            import_cap = get_or_default(row, 'import_capacity')
            ramp_base = get_or_default(row, 'ramplimit')

            if not (pd.notna(grid) and pd.notna(from_node) and pd.notna(to_node)):
                continue  # skip incomplete defs

            def build_row(dir_from, dir_to, cap_value, is_reverse: bool):
                out = {
                    'grid': grid,
                    'from_node': dir_from,
                    'to_node': dir_to,
                    'transferCap': cap_value,
                }
                for p in param_gnn:
                    if p == 'transferCap':
                        continue
                    if p == 'rampLimit':
                        if not is_reverse:
                            out[p] = ramp_base or 0
                        else:
                            out[p] = (ramp_base * (export_cap / import_cap)) if (import_cap and export_cap) else 0
                    else:
                        out[p] = get_or_default(row, p)
                return out

            # left-to-right (export)
            rows.append(build_row(from_node, to_node, export_cap, is_reverse=False))

            # right-to-left (import)
            rows.append(build_row(to_node, from_node, import_cap, is_reverse=True))

        # construct p_gnn
        final_cols = dimensions + param_gnn
        p_gnn = pd.DataFrame(rows, columns=final_cols)
        p_gnn = utils.fill_numeric_na(utils.standardize_df_dtypes(p_gnn))

        # sort by grid, from_node, to_node
        p_gnn.sort_values(
            by=['grid', 'from_node', 'to_node'],
            key=lambda col: col.str.lower() if col.dtype == 'object' else col,
            inplace=True
        )

        # add fake multi-index
        p_gnn = self.create_fake_MultiIndex(p_gnn, dimensions)

        return p_gnn


    def create_p_gn(
        self, 
        p_gnu_io_flat: pd.DataFrame, 
        df_fueldata: pd.DataFrame,
        df_demanddata: pd.DataFrame, 
        df_storagedata: pd.DataFrame,
        ts_storage_limits: dict[str, pd.DataFrame], 
        ts_domain_pairs: dict[str, list]
        ) -> pd.DataFrame:
        """
        Creates a new DataFrame p_gn with specified dimension and parameter columns

        Collects (grid, node) pairs from df_storagedata, p_gnn, and p_gnu
        
        classifies each gn based on a range of tests
            - if node has any fuel data in df_fueldata -> price node
            - if node is not price node -> balance node
            - classifying node as a storage nodes if it passes any of the following three tests
                1. if gn has upwardLimit, downwardLimit, or reference in df_storagedata -> storage node
                2. if node is in ts_storage_limits
                3. if 'upperLimitCapacityRatio' is defined to any (grid, node) -> storage node
                4. if gn is in p_gnu with both 'input' and 'output' roles and grid does not appear in df_demanddata

        Removes zeroes
        """      
        # dimension and parameter columns
        dimensions = ['grid', 'node']
        param_gn = self.PARAM_GN
        
        # List to collect the new rows 
        rows = []

        # --- Collect grid-node pairs ---
        # Get grid_node from ts_domain_pairs and convert to DataFrame
        if 'grid_node' in ts_domain_pairs:
            ts_grid_node = pd.DataFrame(ts_domain_pairs['grid_node'], columns=['grid', 'node'])
        else:
            ts_grid_node = pd.DataFrame(columns=['grid', 'node'])

        # Extract gn pairs from df_demanddata
        if not df_demanddata.empty:
            pairs_df_demanddata = df_demanddata[['grid', 'node']]
        else:
            pairs_df_demanddata = pd.DataFrame(columns=['grid', 'node'])

        # Extract gn pairs from df_storagedata
        if not df_storagedata.empty:
            pairs_df_storagedata = df_storagedata[['grid', 'node']]
        else:
            pairs_df_storagedata = pd.DataFrame(columns=['grid', 'node'])

        # Extract gn pairs from p_gnu_io_flat
        if not p_gnu_io_flat.empty:
            pairs_gnu = p_gnu_io_flat[['grid', 'node']]
        else:
            pairs_gnu = pd.DataFrame(columns=['grid', 'node'])

        # Concatenate and drop duplicates
        parts = [ts_grid_node, pairs_df_demanddata, pairs_df_storagedata, pairs_gnu]
        parts = [p for p in parts if not p.empty]
        unique_gn_pairs = (
            pd.concat(parts, ignore_index=True).drop_duplicates(ignore_index=True)
            if parts else pd.DataFrame(columns=['grid', 'node'])
        )

        # --- Preprocess data for the loop ---

        # Grids in df_demanddata for quick membership test
        if not df_demanddata.empty:
            demand_grids = {str(g).lower() for g in df_demanddata['grid'].dropna().unique()}
        else:
            demand_grids = set()

        # Build a lower-case column map for df_storagedata to support case-insensitive lookup
        if not df_storagedata.empty:
            storage_colmap = {c.lower(): c for c in df_storagedata.columns}
            storage_cols_lower = set(storage_colmap.keys())
        else:
            storage_colmap = {}
            storage_cols_lower = set()

        # Initialize an empty set for storage nodes
        ts_storage_nodes = set()   
        # Iterate through the filtered DataFrames
        for key, df in ts_storage_limits.items():
            # Verify that the DataFrame has the required columns
            if 'node' in df.columns and 'param_gnBoundaryTypes' in df.columns:
                # Add all nodes to the set
                ts_storage_nodes.update(df['node'].unique())

        # --- Process each (grid, node) pair ---
        for _, row in unique_gn_pairs.iterrows():
            grid = row['grid']
            node = row['node']

            # Subset of storagedata for this (grid,node)
            if not df_storagedata.empty:
                node_storage_data = df_storagedata[(df_storagedata['grid'] == grid) & (df_storagedata['node'] == node)]
                node_storage_data = utils.fill_numeric_na(utils.standardize_df_dtypes(node_storage_data))
            else:
                node_storage_data = pd.DataFrame()

            # ---- Basic classifications ----
            isActive = node_storage_data['isActive'].iloc[0] if 'isActive' in node_storage_data.columns else None
            usePrice = node_storage_data['usePrice'].iloc[0] if 'usePrice' in node_storage_data.columns else None
            nodeBalance = node_storage_data['nodeBalance'].iloc[0] if 'nodeBalance' in node_storage_data.columns else None
            energyStoredPerUnitOfState = node_storage_data['energyStoredPerUnitOfState'].iloc[0] if 'energyStoredPerUnitOfState' in node_storage_data.columns else None

            if usePrice == 1 and nodeBalance == 1:
                self.logger.log_status(f"Storage data for (grid, node):({grid}, {node}) has 'usePrice'=1 and 'nodeBalance'=1, check the data.", level="warn")

            if usePrice == 1 and energyStoredPerUnitOfState == 1:
                self.logger.log_status(f"Storage data for (grid, node):({grid}, {node}) has 'usePrice'=1 and 'energyStoredPerUnitOfState'=1, check the data.", level="warn")    

            # --- Check if price node ---

            # usePrice if any fuel record for this grid
            if not usePrice and not df_fueldata.empty:
                usePrice = not df_fueldata[df_fueldata['grid'] == grid].empty
                usePrice = 1 if usePrice else 0

            # --- Check if balance node ---

            # if demand node
            if nodeBalance == None and grid in demand_grids:
                nodeBalance = 1

            # if has any data in node_storage_data
            if nodeBalance == None and not node_storage_data.empty:
                nodeBalance = 1

            # --- Check if storage node ---

            # 1) upwardLimit / downwardLimit / reference present and > 0
            cols_to_check = ['upwardLimit', 'downwardLimit', 'reference']
            existing_lower = [c.lower() for c in cols_to_check if c.lower() in storage_cols_lower]
            if not usePrice and existing_lower and not node_storage_data.empty:
                for low in existing_lower:
                    real_col = storage_colmap[low]
                    numeric_col = pd.to_numeric(node_storage_data[real_col], errors='coerce')
                    if (numeric_col > 0).any():
                        energyStoredPerUnitOfState = 1
                        break

            # 2) node in ts_storage_limits
            if not usePrice and not energyStoredPerUnitOfState and node in ts_storage_nodes:
                energyStoredPerUnitOfState = 1

            # 3) upperLimitCapacityRatio defined (non-null & non-zero) in p_gnu_io_flat
            if not usePrice and not energyStoredPerUnitOfState and not p_gnu_io_flat.empty and 'upperLimitCapacityRatio' in p_gnu_io_flat.columns:
                subset_p_gnu_io = p_gnu_io_flat[(p_gnu_io_flat['grid'] == grid) & (p_gnu_io_flat['node'] == node)]
                if not subset_p_gnu_io.empty:
                    energyStoredPerUnitOfState = ((subset_p_gnu_io['upperLimitCapacityRatio'].notnull()) &
                                  (subset_p_gnu_io['upperLimitCapacityRatio'] != 0)).any()
                    energyStoredPerUnitOfState = 1 if energyStoredPerUnitOfState else 0
            else:
                subset_p_gnu_io = pd.DataFrame()


            # ---- Build the data row ----            
            # Derivatice flags
            if energyStoredPerUnitOfState and usePrice == None: usePrice = 0 
            if energyStoredPerUnitOfState and nodeBalance == None: nodeBalance = 1            


            # Build row with options used for every gn
            row_dict = {
                'grid':                       grid,
                'node':                       node,
                'isActive':                   isActive,
                'usePrice':                   usePrice,
                'nodeBalance':                nodeBalance,
                'energyStoredPerUnitOfState': energyStoredPerUnitOfState
            }

            # Add optional params if present in storagedata (case-insensitive)
            if not node_storage_data.empty:
                for key in (k for k in param_gn if k not in row_dict):
                    low = key.lower()
                    if low in storage_colmap:
                        real_col = storage_colmap[low]
                        val = node_storage_data[real_col].iloc[0]
                        if val is not None:
                            row_dict[key] = val
            rows.append(row_dict)

        
        # Build p_gn
        final_cols = dimensions + param_gn
        p_gn = pd.DataFrame(rows, columns=final_cols)
        p_gn = utils.fill_numeric_na(utils.standardize_df_dtypes(p_gn))
        protected_gn = {'grid', 'node', 'usePrice', 'nodeBalance', 'energyStoredPerUnitOfState'}
        p_gn = p_gn.drop(columns=[col for col in p_gn.columns
                                   if utils.is_col_empty(p_gn[col]) and col not in protected_gn])
        
        # Sort by grid, node in a case-insensitive manner.
        p_gn.sort_values(by=['grid', 'node'], 
                            key=lambda col: col.str.lower() if col.dtype == 'object' else col,
                            inplace=True)

        # Remove zeroes
        p_gn = p_gn.mask(p_gn == 0) 

        # Create the fake multi-index
        p_gn = self.create_fake_MultiIndex(p_gn, dimensions)

        return p_gn


# ------------------------------------------------------
# Functions to create unittype based input tables: 
# unitUnittype, flowUnit, p_unit, effLevelGroupUnit
# ------------------------------------------------------

    def create_unitUnittype(
        self,
        df_unitdata: pd.DataFrame,
        active_units  # array-like of unit names
        ) -> pd.DataFrame:
        # return empty dataframe if no input data
        if df_unitdata.empty:
            return pd.DataFrame(columns=['unit', 'unittype'])
        if 'unittype' not in df_unitdata.columns:
            return pd.DataFrame(columns=['unit', 'unittype'])

        # Get unit-unittype pairs directly from df_unitdata for active units
        unitUnittype = (
            df_unitdata.loc[df_unitdata['unit'].isin(active_units), ['unit', 'unittype']]
            .dropna(subset=['unit', 'unittype'])
            .drop_duplicates(subset=['unit'])
            .reset_index(drop=True)
        )
        return unitUnittype
    

    def create_flowUnit(
        self,
        df_unitdata: pd.DataFrame,
        unitUnittype: pd.DataFrame
        ) -> pd.DataFrame:
        # return empty dataframe if no input data
        if unitUnittype.empty:
            return pd.DataFrame()

        # 'flow' is now directly in the merged df_unitdata.
        if 'flow' not in df_unitdata.columns:
            return pd.DataFrame(columns=['flow', 'unit'])

        # Keep only active units (those that made it through create_p_gnu_io)
        active_units = unitUnittype[['unit']]
        merged = active_units.merge(df_unitdata[['unit', 'flow']], on='unit', how='inner')
        # Exclude rows without a flow value ('' after fill_all_na, or explicit non-empty)
        flowUnit = merged[merged['flow'].notna() & (merged['flow'] != '')][['flow', 'unit']]
        return flowUnit


    def create_p_unit(
        self,
        unitUnittype: pd.DataFrame,
        df_unitdata: pd.DataFrame
        ) -> pd.DataFrame:
        """
        Create the `p_unit` DataFrame for Backbone model input.

        Constructs a parameter table for each unit in `unitUnittype` using the
        merged df_unitdata, which already incorporates type-level defaults via
        merge_unittypedata_into_unitdata().  No separate df_unittypedata lookup
        is needed.

        Parameter values follow this priority (resolved before this method runs):
            1. df_unitdata unit-specific values
            2. Type-level defaults (merged into df_unitdata in the pipeline)
            3. Non-zero parameter defaults (applied in the pipeline)
            4. 0 / def_value fallback in get_param_value

        Enforces the Backbone constraint:
            minShutdownHours <= startWarmAfterXhours <= startColdAfterXhours

        Parameters
        ----------
        unitUnittype : pd.DataFrame
            DataFrame linking individual units to their unittype.
            Must include columns ['unit', 'unittype'].
        df_unitdata : pd.DataFrame
            Merged unit data (unit-specific + type defaults).
            Must include column 'unit'.

        Returns
        -------
        pd.DataFrame
            Finalized `p_unit` DataFrame — one row per unit, sorted by unit name,
            with a fake MultiIndex applied.
        """
        # Dimension column names.
        dimensions = ['unit']

        param_unit = self.PARAM_UNIT
        # List to collect new rows.
        rows = []

        # Process each row in unitUnittype.
        for _, u_row in unitUnittype.iterrows():
            unit = str(u_row['unit'])

            # Case-insensitive matching for unit in the merged df_unitdata.
            unit_matches = df_unitdata[df_unitdata['unit'].str.lower() == unit.lower()]
            if unit_matches.empty:
                self.logger.log_status(f"No unit data found for unit: '{unit}'. "
                            "Not writing the p_unit data, check spelling and files.'",
                            level="warn")
                continue
            unit_row = unit_matches.iloc[0]

            # Pre-fetch minShutdownHours.
            # _ensure_numeric_dtypes guarantees param_unit columns are Float64 with no
            # NA and lowercase names, so a direct Series lookup is sufficient.
            min_shutdown = unit_row.get('minshutdownhours', 0)

            # Start building the row data with the unit column.
            row = {'unit': unit}

            # Loop through the parameters defined in param_unit.
            for param in param_unit:
                # For startColdAfterXhours, compute the maximum of min_shutdown and the fetched value.
                # In Backbone, Units must have minShutdownHours <= startWarmAfterXhours <= startColdAfterXhours
                if param == 'startColdAfterXhours':
                    startColdAfterXhours = unit_row.get(param.lower(), 0)
                    row[param] = max(min_shutdown, startColdAfterXhours)
                    continue

                row[param] = unit_row.get(param.lower(), 0)

            rows.append(row)

        # Build p_unit
        final_cols = dimensions + param_unit
        p_unit = pd.DataFrame(rows, columns=final_cols)
        p_unit = utils.fill_numeric_na(utils.standardize_df_dtypes(p_unit))

        # Sort p_unit by the 'unit' column in a case-insensitive manner.
        p_unit.sort_values(
            by=['unit'],
            key=lambda col: col.str.lower() if col.dtype == 'object' else col,
            inplace=True
        )

        # Apply fake MultiIndex on dimensions
        p_unit = self.create_fake_MultiIndex(p_unit, dimensions)

        return p_unit
    

    def create_effLevelGroupUnit(
        self,
        df_unitdata: pd.DataFrame,
        unitUnittype: pd.DataFrame
        ) -> pd.DataFrame:
        # List to accumulate new rows
        rows = []

        # Iterate over each row in unitUnittype
        for _, u_row in unitUnittype.iterrows():
            unit = u_row['unit']

            # Look up the unit row in the merged df_unitdata.
            unit_matches = df_unitdata[df_unitdata['unit'].str.lower() == str(unit).lower()]
            if unit_matches.empty:
                continue
            unit_row = unit_matches.iloc[0]

            # LP/MIP value (column name lowercased after normalize_dataframe).
            lp_mip = unit_row.get('lp/mip', '') if 'lp/mip' in unit_row.index else ''
            if not isinstance(lp_mip, str):
                lp_mip = ''

            if lp_mip in ['LP', 'MIP']:
                # effLevel1 = MIP/LP
                rows.append({
                    'effLevel': 'level1',
                    'effSelector': f'directOn{lp_mip}',
                    'unit': unit
                })
                # effLevel2-3 = LP
                for i in range(2, 4):
                    rows.append({
                        'effLevel': f'level{i}',
                        'effSelector': 'directOnLP',
                        'unit': unit
                    })

        # Create a new DataFrame from the list of rows with the desired columns
        effLevelGroupUnit = pd.DataFrame(rows, columns=['effLevel', 'effSelector', 'unit'])
        return effLevelGroupUnit


# ------------------------------------------------------
# Functions to create node based input tables: 
# p_gnBoundaryPropertiesForStates, ts_priceChange, p_userconstraint
# ------------------------------------------------------

    def create_p_gnBoundaryPropertiesForStates(
        self, 
        p_gn_flat: pd.DataFrame, 
        df_storagedata: pd.DataFrame, 
        ts_storage_limits: dict[str, pd.DataFrame]
        ) -> pd.DataFrame:
        """
        Creates a DataFrame that defines boundary properties for nodes in an energy grid system.

        This function generates boundary properties for nodes that have balance requirements 
        or energy storage capabilities. It processes both constant values from the storage 
        dataframe and time series data provided in ts_storage_limits.

        Parameters:
        -----------
        p_gn_flat : DataFrame containing node configurations.
            Must include columns: 'grid', 'node', 'nodeBalance', 'energyStoredPerUnitOfState'

        df_storagedata : DataFrame containing storage input data specifications.
            Must include columns: 'grid', 'node', and may include boundary values like
            'upwardlimit', 'downwardlimit', etc.

        ts_storage_limits

        Returns:
        --------
        DataFrame
            A fake-multi-indexed DataFrame with dimensions ['grid', 'node', 'param_gnBoundaryTypes'] 
            and param_gnBoundaryProperties ['useConstant', 'constant', 'useTimeSeries', 'slackCost'].
            Fake multi-index is a compromise between many aspects, see create_fake_MultiIndex()
            Defines boundary constraints for nodes in the energy system.
        """ 
        if p_gn_flat.empty:
            return pd.DataFrame()

        # Define the dimensions and parameters for the fake-multi-indexed output DataFrame
        dimensions = ['grid', 'node', 'param_gnBoundaryTypes']

        # List of supported boundary types to process
        param_gnBoundaryTypes = ['upwardLimit', 'downwardLimit', 
                                 'reference', 'balancePenalty', 
                                 'maxSpill', 'downwardSlack01']
        
        # Properties that will be assigned to each boundary type
        param_gnBoundaryProperties = ['useConstant', 'constant', 'useTimeSeries', 'slackCost']

        # Initialize an empty list to collect all rows for the output DataFrame 
        rows = []   

        # Create a lookup dictionary to quickly find average values for node-boundary type combinations
        # Format: {(node, param_gnBoundaryTypes): average_value}
        ts_node_boundaryTypes = {}

        # Process all time series data frames to populate the lookup dictionary
        for key, df in ts_storage_limits.items():
            # Verify the DataFrame has all required columns before processing
            if all(col in df.columns for col in ['node', 'param_gnBoundaryTypes', 'average_value']):
                # Extract (node, param_gnBoundaryTypes, average_value) tuples and add to dictionary
                for _, row in df[['node', 'param_gnBoundaryTypes', 'average_value']].iterrows():
                    node_boundaryType = (row['node'], row['param_gnBoundaryTypes'])
                    ts_node_boundaryTypes[node_boundaryType] = row['average_value']

        # Process each node in the system that requires balance constraints
        for _, gn_row in p_gn_flat.iterrows():
            # Only process nodes with balance requirements (nodeBalance = 1)
            nodeBalance = gn_row.get('nodeBalance', 0)
            if not utils.is_val_empty(nodeBalance) and nodeBalance == 1:
                grid = gn_row['grid']
                node = gn_row['node']

                # Find the corresponding row in df_storagedata where both grid and node match
                storage_row = None
                if not df_storagedata.empty:
                    mask = (df_storagedata['grid'] == grid) & (df_storagedata['node'] == node)
                    if mask.any():
                        storage_row = df_storagedata[mask].iloc[0]

                # Process each boundary type for this node
                for p_type in param_gnBoundaryTypes:

                    # Check if there's a constant value in storage configuration
                    value = storage_row.get(p_type.lower(), None) if storage_row is not None else None

                    # Time series data takes precedence over constant values
                    if (node, p_type) in ts_node_boundaryTypes:
                        # Create entry for time series-based boundary         
                        row_dict = {
                            'grid':                     grid,
                            'node':                     node,
                            'param_gnBoundaryTypes':    p_type,
                            'useTimeSeries':            1,
                        }
                        rows.append(row_dict)  

                        ## Special case: For downwardLimit with time series, create a slack variable
                        ## to allow minor violations with a smaller penalty
                        #if p_type == 'downwardLimit':
                        #    row_dict = {
                        #        'grid':                     grid,
                        #        'node':                     node,
                        #        'param_gnBoundaryTypes':    'downwardSlack01',
                        #        'useConstant':              1,
                        #        # Scale down the average value and round it
                        #        'constant':                 round(ts_node_boundaryTypes[(node, p_type)]/1000, 0),
                        #        'slackCost':                1000 # Fixed penalty cost for violations
                        #    }
                        #    rows.append(row_dict)      

                    # If no time series but we have a non-zero constant value, use it
                    elif not utils.is_val_empty(value) and isinstance(value, (int, float)) and value > 0:
                        row_dict = {
                            'grid':                     grid,
                            'node':                     node,
                            'param_gnBoundaryTypes':    p_type,
                            'useConstant':              1,
                            'constant':                 value
                        }
                        rows.append(row_dict)
                
            # Additional check for storage nodes
            isStorage = gn_row.get('energyStoredPerUnitOfState', 0)
            if not utils.is_val_empty(isStorage) and (isStorage == 1 or isStorage is True):
                grid = gn_row['grid']
                node = gn_row['node']
                # Ensure all storage nodes have at least an 'Eps' downward limit
                if not any((r['grid'] == grid and 
                            r['node'] == node and 
                            r['param_gnBoundaryTypes'] == 'downwardLimit') for r in rows):
                    row_dict = {
                        'grid':                  grid,
                        'node':                  node,
                        'param_gnBoundaryTypes': 'downwardLimit',
                        'useConstant':           1,
                        'constant':              'Eps'
                    }
                    rows.append(row_dict)


        # Build p_gnBoundaryPropertiesForStates
        final_cols = dimensions + param_gnBoundaryProperties
        p_gnBoundaryPropertiesForStates = pd.DataFrame(rows, columns=final_cols)
        p_gnBoundaryPropertiesForStates = utils.fill_numeric_na(utils.standardize_df_dtypes(p_gnBoundaryPropertiesForStates))

        # Sort by grid, from_node, to_node in a case-insensitive manner.
        p_gnBoundaryPropertiesForStates.sort_values(by=['grid', 'node'], 
                            key=lambda col: col.str.lower() if col.dtype == 'object' else col,
                            inplace=True)

        # Create fake multi-index
        p_gnBoundaryPropertiesForStates = self.create_fake_MultiIndex(p_gnBoundaryPropertiesForStates, dimensions)

        return p_gnBoundaryPropertiesForStates


    def add_storage_starts(
        self, p_gn: pd.DataFrame, 
        p_gnBoundaryPropertiesForStates: pd.DataFrame, 
        p_gnu_io_flat: pd.DataFrame, 
        ts_storage_limits: dict[str, pd.DataFrame]
        ) -> tuple[pd.DataFrame, pd.DataFrame]:
        """
        Adds storage start values to nodes with energy storage capabilities. 
        Converts p_gn and p_gnBoundaryPropertiesForStates to flat versions by removing the fake multi-index.
        Adds p_gn('boundStart') and p_gnBoundaryPropertiesForStates('reference') for storage nodes
        Recreates the fake multi-index

        Parameters:
            p_gn: DataFrame with columns ['grid', 'node'] and possibly 'energyStoredPerUnitOfState'
            p_gnBoundaryPropertiesForStates: DataFrame with columns ['grid', 'node', 'param_gnBoundaryTypes', 'param_gnBoundaryProperties']
            ts_storage_limits

        Returns:
            tuple: (p_gn, p_gnBoundaryPropertiesForStates) with updated values
        """
        if p_gn.empty or p_gnBoundaryPropertiesForStates.empty:
            return(p_gn, p_gnBoundaryPropertiesForStates)
        
        # Create a lookup dictionary to quickly find average values for node-boundary type combinations
        # Format: {(node, param_gnBoundaryTypes): average_value}
        ts_node_boundaryTypes = {}

        # Process all time series data frames to populate the lookup dictionary
        for _, df in ts_storage_limits.items():
            # Verify the DataFrame has all required columns before processing
            if all(col in df.columns for col in ['node', 'param_gnBoundaryTypes', 'average_value']):
                # Extract (node, param_gnBoundaryTypes, average_value) tuples and add to dictionary
                for _, row in df[['node', 'param_gnBoundaryTypes', 'average_value']].iterrows():
                    node_boundaryType = (row['node'], row['param_gnBoundaryTypes'])
                    ts_node_boundaryTypes[node_boundaryType] = row['average_value']

        # Get a flat versions without the fake multi-index
        p_gn_flat = self.drop_fake_MultiIndex(p_gn)
        p_gnBoundaryPropertiesForStates_flat = self.drop_fake_MultiIndex(p_gnBoundaryPropertiesForStates)

        # Identify storage nodes - those where energyStoredPerUnitOfState is 1 or True
        storage_gn = []
        if 'energyStoredPerUnitOfState' in p_gn_flat.columns:
            for _, row in p_gn_flat.iterrows():
                isStorage = row.get('energyStoredPerUnitOfState', 0)
                if not utils.is_val_empty(isStorage) and (isStorage == 1 or isStorage is True):
                    storage_gn.append((row['grid'], row['node']))

        # Add 'boundStart' column to p_gn, initializing with 0
        p_gn_flat['boundStart'] = 0

        # Process each storage node
        for grid, node in storage_gn:

            # 1) calculate start_value based on the timeseries upwardLimit
            start_value = ts_node_boundaryTypes.get((node, 'upwardLimit'), 0)

            # 2) check if there's a constant value in p_gnBoundaryPropertiesForStates_flat
            if start_value == 0:
                # Find rows that match our criteria
                mask = ((p_gnBoundaryPropertiesForStates_flat['grid'] == grid) & 
                       (p_gnBoundaryPropertiesForStates_flat['node'] == node) &
                       (p_gnBoundaryPropertiesForStates_flat['param_gnBoundaryTypes'] == 'upwardLimit'))

                # Check if any matching rows exist and have constant values > 0
                if any(mask) and 'constant' in p_gnBoundaryPropertiesForStates_flat.columns:
                    constant_values = p_gnBoundaryPropertiesForStates_flat.loc[mask, 'constant']
                    if not constant_values.empty and constant_values.iloc[0] > 0:
                        start_value = constant_values.iloc[0]

            # 3) calculate maximum storage based on p_gnu_io('upperLimitCapacityRatio')
            if start_value == 0 and not p_gnu_io_flat.empty:
                subset_p_gnu_io = p_gnu_io_flat[(p_gnu_io_flat['grid'] == grid) & 
                                                (p_gnu_io_flat['node'] == node) & 
                                                (p_gnu_io_flat['upperLimitCapacityRatio'] > 0)
                                                ]
                if not subset_p_gnu_io.empty:
                        # Use the subset dataframe and get the first row if there are multiple matches
                        capacity = subset_p_gnu_io['capacity'].iloc[0]
                        upper_limit = subset_p_gnu_io['upperLimitCapacityRatio'].iloc[0]
                        if pd.notna(capacity):
                            start_value = capacity * upper_limit                      

            # Only proceed with adding/updating p_gn and boundary properties if we have a valid start_value
            if pd.notna(start_value) and start_value >= 0:
                # Set boundStart to 1 for storage nodes
                p_gn_flat.loc[(p_gn_flat['grid'] == grid) & (p_gn_flat['node'] == node), 'boundStart'] = 1

                new_constant = round(start_value * 0.7, 0)
                # Create a mask to find the 'reference' row for this grid and node
                ref_mask = (
                    (p_gnBoundaryPropertiesForStates_flat['grid'] == grid) &
                    (p_gnBoundaryPropertiesForStates_flat['node'] == node) &
                    (p_gnBoundaryPropertiesForStates_flat['param_gnBoundaryTypes'] == 'reference')
                )

                if not p_gnBoundaryPropertiesForStates_flat.loc[ref_mask].empty:
                    # If row exists, update the 'constant' value (and useConstant as needed)
                    p_gnBoundaryPropertiesForStates_flat.loc[ref_mask, 'constant'] = new_constant
                    p_gnBoundaryPropertiesForStates_flat.loc[ref_mask, 'useConstant'] = 1
                else:
                    # Create new row since one does not exist yet.
                    new_row = {
                        'grid': grid,
                        'node': node,
                        'param_gnBoundaryTypes': 'reference',
                        'useConstant': 1,
                        'constant': new_constant
                    }
                    new_row_df = pd.DataFrame([new_row])
                    # Use pandas concat instead of append (which is deprecated in newer pandas versions)
                    p_gnBoundaryPropertiesForStates_flat = pd.concat(
                        [p_gnBoundaryPropertiesForStates_flat, new_row_df],
                        ignore_index=True
                    )

        # Standardize dtypes, fill NA
        p_gn_flat = utils.fill_numeric_na(utils.standardize_df_dtypes(p_gn_flat))
        p_gnBoundaryPropertiesForStates = utils.fill_numeric_na(utils.standardize_df_dtypes(p_gnBoundaryPropertiesForStates))
        
        # Sort p_gnBoundaryPropertiesForStates alphabetically by [grid, node] in a case-insensitive manner
        p_gnBoundaryPropertiesForStates_flat = p_gnBoundaryPropertiesForStates_flat.sort_values(
                                                    by=['grid', 'node', 'param_gnBoundaryTypes'], 
                                                    key=lambda x: x.str.lower()
                                                    ).reset_index(drop=True)

        # recreate fake multi-indexes
        p_gn = self.create_fake_MultiIndex(p_gn_flat, ['grid', 'node'])
        p_gnBoundaryPropertiesForStates = self.create_fake_MultiIndex(p_gnBoundaryPropertiesForStates_flat, ['grid', 'node', 'param_gnBoundaryTypes'])

        return (p_gn, p_gnBoundaryPropertiesForStates)


    def create_ts_priceChange(
        self, 
        p_gn_flat: pd.DataFrame, 
        df_fueldata: pd.DataFrame
        ) -> pd.DataFrame:
        # return empty dataframe if no price nodes (empty p_gn) or no price data (empty df_fueldata)
        if p_gn_flat.empty or df_fueldata.empty:
            return pd.DataFrame()

        # Identify the price column in df_fueldata (case-insensitive), return empty dataframe if not found
        price_col = next((col for col in df_fueldata.columns if col.lower() == 'price'), None)
        if price_col is None:
            return pd.DataFrame()

        rows = []
        # Loop through each row in p_gn using the columns: grid, node, and usePrice
        for _, row in p_gn_flat.iterrows():
            grid_value = row['grid']
            node_value = row['node']

            # Retrieve the node_price from df_fueldata where the grid value matches.
            matching_rows = df_fueldata[df_fueldata['grid'] == grid_value]
            if not matching_rows.empty:
                node_price = matching_rows.iloc[0][price_col]

                # Create a dictionary for the new row
                row_dict = {
                    'node': node_value,
                    't': 't000001',
                    'value': node_price                    
                }
                rows.append(row_dict)

        # Create the ts_priceChange DataFrame from the list of row dictionaries
        ts_priceChange = pd.DataFrame(rows)
        ts_priceChange = utils.fill_numeric_na(utils.standardize_df_dtypes(ts_priceChange))
        
        return ts_priceChange


    def create_p_userconstraint(
        self,
        uc_data: pd.DataFrame,
        p_gnu_io_flat: pd.DataFrame,
        mingen_nodes: list[str],
        ) -> pd.DataFrame:
        """
        Creates the parameter DataFrame `p_userConstraint` defining user constraints.

        The function combines:
          1. Predefined user-constraint data from `uc_data` (added directly as-is,
             with case-insensitive column handling).
          2. Dynamically generated user-constraint rules based on `mingen_nodes`,
             linking nodes to units in `p_gnu_io_flat`.

        Parameters
        ----------
        uc_data : pd.DataFrame
            User-defined constraints, possibly with column names in arbitrary case.
            Expected logical columns (case-insensitive):
            ['group', '1st dimension', '2nd dimension', '3rd dimension',
             '4th dimension', 'parameter', 'value'].
        p_gnu_io_flat : pd.DataFrame
            Flattened DataFrame mapping grids, nodes, and units with columns
            ['grid', 'node', 'unit', 'input_output'].
        mingen_nodes : list[str]
            List of node names for which minimum-generation user constraints
            should be created.
        Returns
        -------
        pd.DataFrame
            Combined DataFrame of all user constraints (`p_userConstraint`).
        """
        expected_cols = [
            'group', '1st dimension', '2nd dimension',
            '3rd dimension', '4th dimension', 'parameter', 'value'
        ]

        frames: list[pd.DataFrame] = []

        # ---- Phase 1: add uc_data, case-insensitive column alignment ----
        if uc_data is not None and not uc_data.empty:
            # map lowercased column name -> canonical expected column
            col_map = {c.lower(): c for c in expected_cols}
            rename_dict = {orig: col_map[orig.lower()] for orig in uc_data.columns if orig.lower() in col_map}
            uc_data_renamed = uc_data.rename(columns=rename_dict)

            missing = [c for c in expected_cols if c not in uc_data_renamed.columns]
            if missing:
                self.logger.log_status(f"uc_data missing required columns (after case-insensitive matching): {missing}",
                           level="warn")

            uc_data_aligned = uc_data_renamed[expected_cols]

            # keep only rows that aren't entirely NA to avoid dtype inference warnings later
            uc_data_aligned = uc_data_aligned.dropna(how="all")

            if not uc_data_aligned.empty:
                frames.append(uc_data_aligned)

        # ---- Phase 2: generate rows for mingen_nodes ----
        generated_rows = []
        for node in mingen_nodes:
            row_gnu = p_gnu_io_flat[
                (p_gnu_io_flat['node'] == node) &
                (p_gnu_io_flat['input_output'] == 'input')
            ]
            group_UC = f"UC_{node}"

            for _, r in row_gnu.iterrows():
                generated_rows.append({
                    'group': group_UC,
                    '1st dimension': r['grid'],
                    '2nd dimension': node,
                    '3rd dimension': r['unit'],
                    '4th dimension': "-",
                    'parameter': "v_gen",
                    'value': -1,
                })

            # group-level rows
            generated_rows += [
                {'group': group_UC, '1st dimension': "-", '2nd dimension': "-", '3rd dimension': "-", '4th dimension': "-", 'parameter': "GT",             'value': -1},
                {'group': group_UC, '1st dimension': "userconstraintRHS", '2nd dimension': "-", '3rd dimension': "-", '4th dimension': "-", 'parameter': "ts_groupPolicy", 'value': 1},
                {'group': group_UC, '1st dimension': "-", '2nd dimension': "-", '3rd dimension': "-", '4th dimension': "-", 'parameter': "penalty",        'value': 2000},
            ]

        if generated_rows:
            frames.append(pd.DataFrame.from_records(generated_rows, columns=expected_cols))

        # ---- Finalize without ever concatenating with an empty frame ----
        if frames:
            p_userConstraint = pd.concat(frames, ignore_index=True)
        else:
            p_userConstraint = pd.DataFrame(columns=expected_cols)

        # Standardize dtypes, fill NA
        p_userConstraint = utils.fill_numeric_na(utils.standardize_df_dtypes(p_userConstraint))

        return p_userConstraint



# ------------------------------------------------------
# Functions to create emission based input tables: 
# p_nEmission, ts_emissionPriceChange, 
# ------------------------------------------------------

    def create_p_nEmission(
        self, 
        p_gn_flat: pd.DataFrame, 
        df_fueldata: pd.DataFrame
        ) -> pd.DataFrame:
        """
        Create p_nEmission['node', 'emission', 'value'] emission factors (tEmission / MWh) for each node.

        Parameters   
        p_gn_flat : pandas DataFrame with columns 'grid' and 'node'.
        df_fueldata : pandas DataFrame with column 'grid' and optional columns 'emission_XX' 
            where XX is emission name (e.g., CO2, CH4).
        """
        # Return empty dataframe if no fuel nodes (empty p_gn) or no emission data (empty df_fueldata)
        if p_gn_flat.empty or df_fueldata.empty:
            return pd.DataFrame()

        # Extract emission names from column names. Return empty dataframe if no emissions.
        emission_cols = [col for col in df_fueldata.columns if col.startswith('emission_')]
        if emission_cols is None:
            return pd.DataFrame()
        else:
            emissions = [col.replace('emission_', '') for col in emission_cols]

        # Create grid_emission DataFrame with (grid, emission) combinations
        grid_emission_data = []
        for grid in df_fueldata['grid'].unique():
            grid_row = df_fueldata[df_fueldata['grid'] == grid].iloc[0]
            for col, emission in zip(emission_cols, emissions):
                if col in grid_row:
                    value = grid_row[col]
                    if pd.notna(value) and value > 0:
                        grid_emission_data.append({
                            'grid': grid,
                            'emission': emission,
                            'value': value
                        })

        grid_emission = pd.DataFrame(grid_emission_data)

        # Filter grid_node to only include grids that are in grid_emission
        valid_grids = grid_emission['grid'].unique()
        grid_node = p_gn_flat[p_gn_flat['grid'].isin(valid_grids)][['grid', 'node']]

        # Create p_nEmission by joining grid_node with grid_emission
        p_nEmission_data = []
        for _, row in grid_node.iterrows():
            for emission in emissions:
                grid_emission_row = grid_emission[(grid_emission['grid'] == row['grid']) & 
                                                 (grid_emission['emission'] == emission)]
                if not grid_emission_row.empty:
                    value = grid_emission_row.iloc[0]['value']
                    if pd.notna(value) and value > 0:
                        p_nEmission_data.append({
                            'node': row['node'],
                            'emission': emission,
                            'value': value
                        })

        p_nEmission = pd.DataFrame(p_nEmission_data)
        p_nEmission = utils.fill_numeric_na(utils.standardize_df_dtypes(p_nEmission))

        return p_nEmission


    def create_ts_emissionPriceChange(
        self, 
        df_emissiondata: pd.DataFrame
        ) -> pd.DataFrame:
        """
        Create ts_emissionPriceChange ['emission', 'group', 't', 'value'] DataFrame
        
        Parameters: 
            df_emissiondata : pandas DataFrame with columns 'emission', 'group', and optional 'price'.
        """
        # Return empty dataframe if no emission data (empty df_emissiondata)
        if df_emissiondata.empty:
            return pd.DataFrame()
        
        # Extract emission_group pairs
        emission_group = df_emissiondata[['emission', 'group']].drop_duplicates()

        # Check if price column exists
        has_price = 'price' in df_emissiondata.columns

        # Create ts_emissionPriceChange DataFrame
        ts_emissionPriceChange_data = []
        for _, row in emission_group.iterrows():
            emission = row['emission']
            group = row['group']

            # Get price value if it exists
            price_value = 0
            if has_price:
                emission_row = df_emissiondata[(df_emissiondata['emission'] == emission) & 
                                           (df_emissiondata['group'] == group)]
                if not emission_row.empty and not pd.isna(emission_row.iloc[0]['price']):
                    price_value = emission_row.iloc[0]['price']

            ts_emissionPriceChange_data.append({
                'emission': emission,
                'group': group,
                't': 't000001',
                'value': price_value
            })

        ts_emissionPriceChange = pd.DataFrame(ts_emissionPriceChange_data)
        ts_emissionPriceChange = utils.fill_numeric_na(utils.standardize_df_dtypes(ts_emissionPriceChange))


        return ts_emissionPriceChange


    def create_gnGroup(
        self,
        p_nEmission: pd.DataFrame,
        ts_emissionPriceChange: pd.DataFrame,
        p_gnu_io_flat: pd.DataFrame,
        df_unitdata: pd.DataFrame,
        input_dfs: list[pd.DataFrame] = []
        ) -> pd.DataFrame:
        """
        Build gnGroup['grid', 'node', 'group'] from two sources.

        Source 1 — emission-based lookup (five-step join):
          1. p_nEmission(node, emission)
          2. ts_emissionPriceChange(emission -> group)  [case-insensitive match]
          3. p_gnu_io_flat(node -> grid, unit)
          4. unitUnittype(unit -> unittype)
          5. df_unitdata: unit row must have at least one emission_group*
             column whose value equals the group from step 2.
          A (grid, node, group) row is added for every combination that passes
          all five steps.

        Source 2 — direct rows from input_dfs:
          Any DataFrame in input_dfs that contains 'grid', 'node', and 'group'
          columns contributes its rows directly, without any emission lookup.

        Warnings are logged when:
          - p_nEmission is empty (no emission nodes defined).
          - ts_emissionPriceChange is empty (no emission price data loaded).
          - Some emission types in p_nEmission have no matching entry in
            ts_emissionPriceChange (logged with the unmatched names).
          - df_unitdata has no emission_group* columns at all.

        Duplicates across both sources are dropped before returning.
        """
        # Initialize an empty list to store rows
        rows_list = []

        # Step 1: Process emissions data
        for _, node_emission in p_nEmission.iterrows():
            node = node_emission['node']
            emission = node_emission['emission']

            # Check if emission exists in ts_emissionPriceChange
            if not ts_emissionPriceChange.empty:
                matching_emission = ts_emissionPriceChange[ts_emissionPriceChange['emission'].str.lower() == emission.lower()] 
            else:
                matching_emission = pd.DataFrame()            

            # Skip the rest of step 1 if no matching emission is found
            if matching_emission.empty:
                continue

            # Get the group value from ts_emissionPriceChange
            group = matching_emission['group'].iloc[0]

            # Find matching grid_node_unit tuples
            matching_gnu = p_gnu_io_flat[p_gnu_io_flat['node'] == node]

            for _, grid_node_unit in matching_gnu.iterrows():
                grid = grid_node_unit['grid']
                unit = grid_node_unit['unit']

                # Look up the unit row in the merged df_unitdata.
                unit_matches = df_unitdata[df_unitdata['unit'] == unit]
                if unit_matches.empty:
                    continue
                unit_row = unit_matches.iloc[0]

                # Find if emission exists in any emission_group column.
                emission_group_cols = [col for col in df_unitdata.columns if col.startswith('emission_group')]
                if not emission_group_cols:
                    self.logger.log_status(
                        "df_unitdata has no 'emission_group*' columns and thus,"
                        "None of the units are producing any emissions. Check that the unittype"
                        "source file(s) includes emission_group column(s).",
                        level="warn"
                    )
                for col in emission_group_cols:
                    if col in unit_row.index and unit_row[col] == group:
                        # Create row and add to rows_list
                        row = {'grid': grid, 'node': node, 'group': group}
                        rows_list.append(row)

        # Step 2: Process input DataFrames
        for df in input_dfs:
            # Check if df has required columns
            if all(col in df.columns for col in ['grid', 'node', 'group']):
                for _, row in df.iterrows():
                    rows_list.append({
                        'grid': row['grid'],
                        'node': row['node'],
                        'group': row['group']
                    })

        # Step 3: create the final DataFrame and drop duplicates
        gnGroup = pd.DataFrame(rows_list)
        gnGroup = pd.DataFrame(rows_list).drop_duplicates()

        return gnGroup


# ------------------------------------------------------
# Utility functions
# ------------------------------------------------------

    def compile_domain_df(
        self, 
        values: list, 
        domain: str
        ) -> pd.DataFrame:
        """
        Produce the final single-column domain DataFrame ready to write to the Backbone input Excel.

        Deduplicates case-insensitively (first occurrence wins) and sorts alphabetically.
        This is the last step before output -- call this once all sources have been gathered
        into a flat list.

        Parameters:
        - values: list of domain values (strings) collected from all sources
        - domain: name for the output column

        Returns:
        - pd.DataFrame with one column named `domain`, or empty DataFrame if no values
        """
        if not values:
            return pd.DataFrame()

        domain_mapping = {}
        for d in values:
            if isinstance(d, str):
                lower_d = d.lower()
                if lower_d not in domain_mapping:
                    domain_mapping[lower_d] = d

        if not domain_mapping:
            return pd.DataFrame()

        result = pd.DataFrame({domain: list(domain_mapping.values())})
        result = result.sort_values(by=domain, key=lambda x: x.str.lower()).reset_index(drop=True)
        return result


    def create_fake_MultiIndex(
        self, 
        df: pd.DataFrame, 
        dimensions: list[str]
        ) -> pd.DataFrame:
        """
        Creates a fake MultiIndex by:
        1. Taking an existing DataFrame with single-layer column names
        2. Creating a new first row with empty strings for dimensions and parameter names for parameter columns
        3. Shifting existing data down by one row

        Parameters:
        - df: pandas DataFrame with single-layered column names
        - dimensions: list of column names that are dimension columns

        Returns:
        - DataFrame with fake MultiIndex structure
        """
        # Identify parameter columns (those not in dimensions)
        all_columns = list(df.columns)

        # Create a new DataFrame with the same columns
        df_output = pd.DataFrame(columns=all_columns)

        # Create the first row with empty strings for dimension columns
        # and parameter names for parameter columns
        first_row = []
        for col in all_columns:
            if col in dimensions:
                first_row.append("")
            else:
                first_row.append(col)

        # Add the first row to the DataFrame
        df_output.loc[0] = first_row

        # Reset the index of the input DataFrame and add it to the output
        df_reset = df.reset_index(drop=True)
        df_reset.index = df_reset.index + 1  # Shift indices to start from 1

        # Concatenate the first row with the original data
        df_output = pd.concat([df_output, df_reset], axis=0)
        df_output = utils.standardize_df_dtypes(df_output)


        return df_output
    

    def drop_fake_MultiIndex(
        self, df: pd.DataFrame
        ) -> pd.DataFrame:
        # early exit if empty input
        if df.empty:
            return df

        # Create a copy of the original DataFrame
        df_flat = df.copy()

        # Drop the first row by using its index position 0
        df_flat = df_flat.drop(df_flat.index[0]).reset_index(drop=True)

        return df_flat


# ------------------------------------------------------
# Post-processing
# ------------------------------------------------------

    def add_index_sheet(self) -> None:
        """
        Adds Index sheet to the excel
            * loads preconstructed 'indexSheet.xlsx'
            * picks rows where Symbol is in the sheet names
        """
        # warn if input folder is not defined - missing Index sheet will prevent data loading to Backbone
        if self.input_folder == "":
            self.logger.log_status("Input folder is not defined - Index sheet was not added. This will prevent data loading to Backbone.", level="warn")
            return

        # Construct full path to the index sheet file
        index_path = os.path.join(self.input_folder, 'indexSheet.xlsx')

        # Read the index sheet file (assuming the first row contains headers)
        try:
            df_index = pd.read_excel(index_path, header=0)
        except Exception:
            self.logger.log_status(f"'{index_path}' not found, index sheet was not added to the BB input Excel.", level="warn")
            return

        # Load the output Excel workbook which already has multiple sheets
        wb = load_workbook(self.output_file)
        existing_sheet_names = wb.sheetnames

        # Filter rows: keep only rows where the 'Symbol' exists among the workbook's sheet names
        df_filtered = df_index[df_index['Symbol'].isin(existing_sheet_names)]

        # Create a new sheet named 'index'
        new_sheet = wb.create_sheet(title='index')

        # Write header row (row 1)
        for col_num, header in enumerate(df_index.columns, start=1):
            new_sheet.cell(row=1, column=col_num, value=header)

        # Write the filtered data starting from row 2
        for row_num, row in enumerate(df_filtered.itertuples(index=False, name=None), start=2):
            for col_num, value in enumerate(row, start=1):
                new_sheet.cell(row=row_num, column=col_num, value=value)

        # Move the 'index' sheet to the first position in the workbook
        wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(new_sheet)))

        # Save the updated workbook back to the output file
        wb.save(self.output_file)


    def adjust_excel(self) -> None:
        """
        For each sheet in the Excel file
            * Adjust each column's width.
            * Skip remaining processing if sheet has only 1 row.
            * If A2 is empty, iterate non-empty cells in row 2:
                - Rotate matching cell in row 1 if the length of the cell is more than 6 letters.
                - Centre align columns
                - set the column width to 6
            * Freeze top row
            * Create and apply table formatting
            * Add explanatory texts after (right from) the generated table in case of "fake MultiIndex"

        Note: Empty A2 means the sheet has "fake MultiIndex" used as a compromize between excel and Backbone
        """
        wb = load_workbook(self.output_file)

        for ws in wb.worksheets:
            max_row = ws.max_row
            max_col = ws.max_column

            # Adjust each column's width (based on longest value in column)
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0

                for row_idx in range(1, max_row + 1):
                    value = ws.cell(row=row_idx, column=col_idx).value
                    if value is not None:
                        length = len(str(value))
                        if length > max_length:
                            max_length = length

                ws.column_dimensions[col_letter].width = max_length + 6  # padding


            # Skip remaining processing if sheet has only 1 row
            if ws.max_row == 1:
                continue

            # If A2 is empty, the sheet has "fake MultiIndex" used as a compromize between excel and GDXXRW
            has_fake_multiindex = ws["A2"].value is None
            if has_fake_multiindex:
                # Pre-create alignments to avoid recreating them in loops
                center_align = Alignment(horizontal="center")
                rotated_header_align = Alignment(horizontal="center", textRotation=90)

                # Iterate cells in row 2 if cells are not empty
                for cell in ws[2]:
                    if cell.value is None:
                        continue

                    col_idx = cell.col_idx
                    col_letter = get_column_letter(col_idx)

                    # Rotate matching cell in row 1 if the length of the cell is more than 6 letters.
                    header_cell = ws.cell(row=1, column=col_idx)
                    header_text = str(header_cell.value) if header_cell.value is not None else ""
                    if len(header_text) > 6:
                        header_cell.alignment = rotated_header_align

                    # Centre align column values from row 2 downwards
                    for row_idx in range(2, max_row + 1):
                        ws.cell(row=row_idx, column=col_idx).alignment = center_align

                    # Set the column width to 6 for these rotated / "special" columns
                    ws.column_dimensions[col_letter].width = 6

            # Freeze the top row
            ws.freeze_panes = "A2"

            # Create and apply table formatting
            # Derive table name from sheet name: remove any non-word characters and append _table.
            table_name = re.sub(r'\W+', '_', ws.title) + "_table"
            # Apply Excel table formatting
            last_col_letter = get_column_letter(ws.max_column)
            table_ref = f"A1:{last_col_letter}{ws.max_row}"
            table = Table(displayName=table_name, ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9",
                                   showFirstColumn=False,
                                   showLastColumn=False,
                                   showRowStripes=True,
                                   showColumnStripes=False)
            table.tableStyleInfo = style
            table.headerRowCount = 1
            ws.add_table(table)

            # If fake MultiIndex, add explanatory texts to the right of the table
            if ws["A2"].value is None:
                n = ws.max_column + 2
                ws.cell(row=1, column=n, value='The first row labels are for excel Table headers.')
                ws.cell(row=2, column=n, value='The Second row labels are for GDXXRW converting excel to GDX.')

        # save the adjusted file
        wb.save(self.output_file)



# ------------------------------------------------------
# Pre checks
# ------------------------------------------------------

    def _normalize_unitdata_columns(self) -> None:
        """
        Restore the _output1 suffix on unsuffixed param_gnu columns in df_unitdata.

        normalize_dataframe (pipeline step 7) strips _output1 from numeric columns,
        so a bare column like 'capacity' implicitly represents output1.  This method
        makes that relationship explicit so downstream lookups can use the uniform
        '{param}_{put}' form (e.g. 'capacity_output1') for all connections.

        Params also in PARAM_UNIT (currently only 'isActive') are shared between
        create_p_gnu_io (needs '{param}_output1') and create_p_unit (needs base form).
        Those get a '{col}_output1' copy while the base column is kept.
        """
        _gnu  = {p.lower() for p in self.PARAM_GNU}
        _unit = {p.lower() for p in self.PARAM_UNIT}
        _valid_put_suffixes = {
            f'_{d}{i}' for d in ('input', 'output') for i in range(1, 6)
        }
        _gnu_only   = _gnu - _unit
        _gnu_shared = _gnu & _unit

        df = self.df_unitdata.copy()
        rename_map = {}
        copy_cols  = {}
        for col in list(df.columns):
            col_l = col.lower()
            if any(col_l.endswith(s) for s in _valid_put_suffixes):
                continue  # already has a connection suffix
            if col_l in _gnu_only:
                output1_name = f'{col}_output1'
                if output1_name not in df.columns:
                    rename_map[col] = output1_name
            elif col_l in _gnu_shared:
                output1_name = f'{col}_output1'
                if output1_name not in df.columns:
                    copy_cols[output1_name] = df[col].copy()
        df = df.rename(columns=rename_map)
        for new_col, series in copy_cols.items():
            df[new_col] = series
        self.df_unitdata = df

    def _ensure_numeric_dtypes(self) -> None:
        """
        Cast all known parameter columns to Float64 and fill NA -> 0.

        Operates on df_unitdata, df_transferdata, df_storagedata, df_demanddata.
        Unknown columns (grid_*, node_*, emission_group*, flow, lp/mip, …) are left
        untouched.  param_unit columns with a connection suffix trigger a warning and
        are left untouched (ignored downstream).

        Must be called after _normalize_unitdata_columns() so that param_gnu columns
        already carry explicit connection suffixes.
        """
        _gnu  = {p.lower() for p in self.PARAM_GNU}
        _unit = {p.lower() for p in self.PARAM_UNIT}
        _valid_put_suffixes = {
            f'_{d}{i}' for d in ('input', 'output') for i in range(1, 6)
        }

        # --- df_unitdata: PARAM_GNU + PARAM_UNIT ---
        df = self.df_unitdata.copy()
        for col in df.columns:
            col_l = col.lower()
            base, put = col_l, None
            for suffix in _valid_put_suffixes:
                if col_l.endswith(suffix):
                    base = col_l[:-len(suffix)]
                    put = suffix[1:]
                    break
            if base in _gnu:
                df[col] = pd.to_numeric(df[col], errors='coerce').astype('Float64').fillna(0)
            elif base in _unit:
                if put is not None:
                    self.logger.log_status(
                        f"Unit data column '{col}': '{base}' is a unit-level parameter "
                        f"(param_unit) and cannot be connection-specific. "
                        f"The '_{put}' suffix is not valid here — column will be ignored.",
                        level="warn"
                    )
                else:
                    df[col] = pd.to_numeric(df[col], errors='coerce').astype('Float64').fillna(0)
        self.df_unitdata = df

        # --- df_transferdata: PARAM_GNN ---
        _gnn = {p.lower() for p in self.PARAM_GNN}
        df = self.df_transferdata.copy()
        for col in df.columns:
            if col.lower() in _gnn:
                df[col] = pd.to_numeric(df[col], errors='coerce').astype('Float64').fillna(0)
        self.df_transferdata = df

        # --- df_storagedata + df_demanddata: PARAM_GN ---
        _gn = {p.lower() for p in self.PARAM_GN}
        for attr in ('df_storagedata', 'df_demanddata'):
            df = getattr(self, attr).copy()
            for col in df.columns:
                if col.lower() in _gn:
                    df[col] = pd.to_numeric(df[col], errors='coerce').astype('Float64').fillna(0)
            setattr(self, attr, df)



# ------------------------------------------------------
# Main entry point for the script
# ------------------------------------------------------

    def run(self) -> None:

        # --- Pre-checks ---

        # Check if the Excel file is locked (e.g. open in Excel) before proceeding
        if os.path.exists(self.output_file):
            try:
                with open(self.output_file, 'a'):
                    pass
            except OSError:
                self.logger.log_status(
                    f"The Backbone input excel file '{self.output_file}' is currently open. Please close it and rerun the code.",
                    level="warn"
                )
                return

        # Restore explicit _output1 suffix on unsuffixed param_gnu columns in df_unitdata.
        self._normalize_unitdata_columns()
        # Enforce Float64 on all known numeric columns; warn on invalid column patterns.
        self._ensure_numeric_dtypes()

        # --- Convert input data tables to DataFrames ---

        # Create p_gnu_io
        p_gnu_io = self.create_p_gnu_io(self.df_unitdata)

        # Create flat version for easier use in other functions
        p_gnu_io_flat = self.drop_fake_MultiIndex(p_gnu_io)

        # unit, unittype domain tables - derived from df_unitdata
        active_units = p_gnu_io_flat['unit'].dropna().unique()
        unit = self.compile_domain_df(active_units.tolist(), 'unit')
        unittype_vals = (
            self.df_unitdata
            .loc[self.df_unitdata['unit'].isin(active_units), 'unittype']
            .dropna().unique().tolist()
        )
        unittype = self.compile_domain_df(unittype_vals, 'unittype')

        # unitUnittype - unit-unittype pairs from df_unitdata for active units
        unitUnittype = self.create_unitUnittype(self.df_unitdata, active_units)
        p_unit = self.create_p_unit(unitUnittype, self.df_unitdata)
        # remove zeroes
        p_unit = p_unit.mask(p_unit == 0)

        # Remove units without capacity or investment parameters
        p_gnu_io, p_unit = self.drop_redundant_units(p_gnu_io, p_unit)
        p_gnu_io_flat = self.drop_fake_MultiIndex(p_gnu_io)
        
        # Calculate missing input or output capacities in p_gnu_io
        p_gnu_io = self.fill_capacities(p_gnu_io, p_unit)
        p_gnu_io_flat = self.drop_fake_MultiIndex(p_gnu_io)

        # Create remaining unit related tables
        flowUnit = self.create_flowUnit(self.df_unitdata, unitUnittype)
        effLevelGroupUnit = self.create_effLevelGroupUnit(self.df_unitdata, unitUnittype)

        # p_gnn
        p_gnn = self.create_p_gnn(self.df_transferdata)

        # Create flat version for easier use in other functions
        p_gnn_flat = self.drop_fake_MultiIndex(p_gnn)

        # p_gn
        p_gn = self.create_p_gn(p_gnu_io_flat, self.df_fueldata, self.df_demanddata, 
                                self.df_storagedata, self.ts_storage_limits, self.ts_domain_pairs)
                  
        # Create flat version for easier use in other functions
        p_gn_flat = self.drop_fake_MultiIndex(p_gn)

        # node based input tables
        p_gnBoundaryPropertiesForStates = self.create_p_gnBoundaryPropertiesForStates(p_gn_flat, 
                                                                                      self.df_storagedata, 
                                                                                      self.ts_storage_limits)
        # remove zeroes
        p_gnBoundaryPropertiesForStates = p_gnBoundaryPropertiesForStates.mask(p_gnBoundaryPropertiesForStates == 0)        
        ts_priceChange = self.create_ts_priceChange(p_gn_flat, self.df_fueldata)
        p_userconstraint = self.create_p_userconstraint(self.df_userconstraintdata,
                                                        p_gnu_io_flat,
                                                        self.mingen_nodes)

        # add storage start levels to p_gn and p_gnBoundaryPropertiesForStates
        (p_gn, p_gnBoundaryPropertiesForStates) = self.add_storage_starts(p_gn, p_gnBoundaryPropertiesForStates, 
                                                                          p_gnu_io_flat, self.ts_storage_limits)
        p_gn_flat = self.drop_fake_MultiIndex(p_gn)

        # emission based input tables
        p_nEmission = self.create_p_nEmission(p_gn_flat, self.df_fueldata)
        ts_emissionPriceChange = self.create_ts_emissionPriceChange(self.df_emissiondata)

        # group sets
        gnGroup = self.create_gnGroup(p_nEmission, ts_emissionPriceChange, p_gnu_io_flat,
                                      self.df_unitdata)


        # --- Compile remaining domains ---

        # grid
        grid_vals = []
        for df in [p_gnu_io_flat, p_gnn_flat, p_gn_flat]:
            if 'grid' in df.columns:
                grid_vals.extend(df['grid'].dropna().tolist())
        if self.ts_domains is not None and 'grid' in self.ts_domains:
            grid_vals.extend(self.ts_domains['grid'])
        grid = self.compile_domain_df(grid_vals, 'grid')

        # node
        node_vals = []
        for df in [p_gnu_io_flat, p_gn_flat]:
            if 'node' in df.columns:
                node_vals.extend(df['node'].dropna().tolist())
        for col in ['from_node', 'to_node']:
            if col in p_gnn_flat.columns:
                node_vals.extend(p_gnn_flat[col].dropna().tolist())
        if self.ts_domains is not None and 'node' in self.ts_domains:
            node_vals.extend(self.ts_domains['node'])
        node = self.compile_domain_df(node_vals, 'node')

        # flow
        flow_vals = flowUnit['flow'].dropna().tolist() if 'flow' in flowUnit.columns else []
        if self.ts_domains is not None and 'flow' in self.ts_domains:
            flow_vals.extend(self.ts_domains['flow'])
        flow = self.compile_domain_df(flow_vals, 'flow')

        # group
        group_vals = []
        for df in [p_userconstraint, ts_emissionPriceChange, gnGroup]:
            if 'group' in df.columns:
                group_vals.extend(df['group'].dropna().tolist())
        if self.ts_domains is not None and 'group' in self.ts_domains:
            group_vals.extend(self.ts_domains['group'])
        group = self.compile_domain_df(group_vals, 'group')

        # emission
        emission_vals = []
        for df in [ts_emissionPriceChange, p_nEmission]:
            if 'emission' in df.columns:
                emission_vals.extend(df['emission'].dropna().tolist())
        emission = self.compile_domain_df(emission_vals, 'emission')

        # restype
        restype = pd.DataFrame()

        # --- scenario tags to an excel sheet ---

        # Alternative columns are added only when present; column names follow the
        # pattern alternative, alternative2, alternative3, alternative4.
        _alt_col_names = ['alternative', 'alternative2', 'alternative3', 'alternative4']
        _n_alts = len(self.scen_tags) - 2
        scen_tags_df = pd.DataFrame([self.scen_tags], columns=['scenario', 'year'] + _alt_col_names[:_n_alts])


        # --- Write DataFrames to excel ---

        # Write DataFrames to different sheets of the merged Excel file
        with pd.ExcelWriter(self.output_file) as writer:
            # scenario tags
            scen_tags_df.to_excel(writer, sheet_name='add_scen_tags', index=False)  

            # node based input tables
            grid.to_excel(writer, sheet_name='grid', index=False)       
            node.to_excel(writer, sheet_name='node', index=False) 
            p_gn.to_excel(writer, sheet_name='p_gn', index=False) 
            p_gnBoundaryPropertiesForStates.to_excel(writer, sheet_name='p_gnBoundaryPropertiesForStates', index=False)        
            ts_priceChange.to_excel(writer, sheet_name='ts_priceChange', index=False)   

            # transfer input tables
            p_gnn.to_excel(writer, sheet_name='p_gnn', index=False)

            # unit input tables
            unit.to_excel(writer, sheet_name='unit', index=False)   
            unittype.to_excel(writer, sheet_name='unittype', index=False)             
            unitUnittype.to_excel(writer, sheet_name='unitUnittype', index=False)     
            flowUnit.to_excel(writer, sheet_name='flowUnit', index=False)            
            effLevelGroupUnit.to_excel(writer, sheet_name='effLevelGroupUnit', index=False)  
            p_gnu_io.to_excel(writer, sheet_name='p_gnu_io', index=False)
            p_unit.to_excel(writer, sheet_name='p_unit', index=False)       
            p_userconstraint.to_excel(writer, sheet_name='p_userconstraint', index=False)     

            # emission based input tables
            p_nEmission.to_excel(writer, sheet_name='p_nEmission', index=False)   
            ts_emissionPriceChange.to_excel(writer, sheet_name='ts_emissionPriceChange', index=False)  

            # group sets
            gnGroup.to_excel(writer, sheet_name='gnGroup', index=False) 

            # remaining domains
            group.to_excel(writer, sheet_name='group', index=False)              
            flow.to_excel(writer, sheet_name='flow', index=False)                                   
            emission.to_excel(writer, sheet_name='emission', index=False)                                   
            restype.to_excel(writer, sheet_name='restype', index=False)    

        # --- Finishing touches ---

        # Apply the adjustments on the Excel file
        self.add_index_sheet()
        self.adjust_excel()

        self.logger.log_status(f"Input excel for Backbone written to '{self.output_file}'", level="info")
        self.bb_excel_succesfully_built = True
