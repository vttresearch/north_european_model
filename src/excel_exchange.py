import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import re
from pathlib import Path
from typing import Union, List, Sequence
from src.utils import log_status


def read_input_excels(
    input_folder: Union[Path, str],
    files: Sequence[str],
    sheet_name_prefix: str,
    logs: List[str],
    *,
    is_mandatory: bool = True,
    drop_note_columns: bool = True,
    add_source_cols: bool = True,
) -> List[pd.DataFrame]:
    """
    Read Excel sheets whose names start with `sheet_name_prefix` (case-insensitive) from
    each file in `files` and return one DataFrame per matched sheet.

    Cleaning steps
    --------------
    - Drop columns whose header is empty/whitespace or auto-generated as 'Unnamed: x'.
    - Optionally drop any column named 'note' (any casing).
    - Truncate the DataFrame at the first fully empty row (drop that row and all below).
      An "empty" cell is NA or a string consisting only of whitespace.

    Parameters
    ----------
    input_folder : Union[Path, str]
        Directory containing the Excel files.
    files : Sequence[str]
        Filenames (relative to `input_folder`) to scan.
    sheet_name_prefix : str
        Case-insensitive prefix used to select sheets within each workbook.
    logs : List[str]
        Log accumulator passed to `log_status`.
    is_mandatory : bool, default True
        If True, missing sheets/files emit warnings; otherwise informational logs.
    drop_note_columns : bool, default True
        If True, drop any column named 'note' (any casing).
    add_source_cols : bool, default True
        If True, append '_source_file' and '_source_sheet'.

    Returns
    -------
    List[pd.DataFrame]
        One DataFrame per matched sheet. Returns [] if nothing was loaded.
    """
    dataframes: List[pd.DataFrame] = []
    input_folder = str(input_folder)

    log_status(
        f"Reading Excel files for '{sheet_name_prefix}': {len(files)} file(s) ...",
        logs, level=None
    )

    for file_name in files:
        file_path = os.path.join(input_folder, file_name)
        try:
            xls = pd.ExcelFile(file_path)
        except Exception as e:
            log_status(f"Unable to open {file_path}: {e}", logs, level="warn")
            continue

        # Match sheets by prefix (case-insensitive)
        matched = [s for s in xls.sheet_names if s.lower().startswith(sheet_name_prefix.lower())]
        if not matched:
            level = "warn" if is_mandatory else "info"
            log_status(
                f"No sheets starting with '{sheet_name_prefix}' in '{file_name}'.",
                logs, level=level
            )
            continue

        for sheet in matched:
            try:
                df = pd.read_excel(xls, sheet_name=sheet, header=0)
            except Exception as e:
                log_status(
                    f"Failed reading sheet '{sheet}' in '{file_name}': {e}",
                    logs, level="warn"
                )
                continue

            # --- Drop columns with empty titles (incl. 'Unnamed: x') ---
            # Consider empty if None, whitespace-only, or header starts with 'unnamed:' (case-insensitive).
            empty_title_cols = [
                c for c in df.columns
                if c is None
                or (isinstance(c, str) and (c.strip() == "" or c.lower().startswith("unnamed:")))
            ]
            if empty_title_cols:
                df = df.drop(columns=empty_title_cols, errors="ignore")

            # Optionally drop 'note' columns (any casing)
            if drop_note_columns:
                note_cols = [c for c in df.columns if isinstance(c, str) and c.lower() == "note"]
                if note_cols:
                    df = df.drop(columns=note_cols, errors="ignore")

            # --- Truncate at the first fully empty row ---
            # Treat empty strings/whitespace as NA, then find first all-NA row.
            _tmp = df.replace(r"^\s*$", pd.NA, regex=True)
            row_is_empty = _tmp.isna().all(axis=1)

            if row_is_empty.any():
                first_empty_pos = row_is_empty.values.argmax()  # position (0-based) of first True
                df = df.iloc[:first_empty_pos]  # drop the empty row and everything below

            # Optionally add provenance columns
            if add_source_cols:
                df = df.copy()
                df["_source_file"] = file_name
                df["_source_sheet"] = sheet

            dataframes.append(df)

    if not dataframes:
        level = "warn" if is_mandatory else "info"
        log_status(
            f"No dataframes loaded for prefix '{sheet_name_prefix}'. Returning empty list.",
            logs, level=level
        )
        return []

    return dataframes



def add_index_sheet(input_folder, output_file):
    """
    Adds Index sheet to the excel
        * loads preconstructed 'indexSheet.xlsx'
        * picks rows where Symbol is in the sheet names
    """
    # skip processing if input folder is not defined
    if input_folder == "":
        return

    # Construct full path to the index sheet file
    index_path = os.path.join(input_folder, 'indexSheet.xlsx')

    # Read the index sheet file (assuming the first row contains headers)
    try:
        df_index = pd.read_excel(index_path, header=0)
    except:
        log_status(f"Warning, '{index_path}' not found and index sheet was not added to the BB input excel", self.builder_logs, level="warn")
        return

    # Load the output Excel workbook which already has multiple sheets
    wb = load_workbook(output_file)
    existing_sheet_names = wb.sheetnames

    # Filter rows: keep only rows where the 'Symbol' exists among the workbook's sheet names
    df_filtered = df_index[df_index['Symbol'].isin(existing_sheet_names)]

    # Create a new sheet named 'index'
    new_sheet = wb.create_sheet(title='index')

    # Write header row (row 1)
    for col_num, header in enumerate(df_index.columns, start=1):
        new_sheet.cell(row=1, column=col_num, value=header)

    # Write the filtered data starting from row 2
    for row_num, row in enumerate(df_filtered.itertuples(index=False, name=None), start=2):
        for col_num, value in enumerate(row, start=1):
            new_sheet.cell(row=row_num, column=col_num, value=value)

    # Move the 'index' sheet to the first position in the workbook
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(new_sheet)))

    # Save the updated workbook back to the output file
    wb.save(output_file)


def adjust_excel(output_file):
    """
    For each sheet in the Excel file
        * Adjust each column's width.
        * Skip remaining processing if sheet has only 1 row.
        * If A2 is empty, iterate non-empty cells in row 2:
                - Rotate matching cell in row 1 if the length of the cell is more than 6 letters.
                - Centre align columns
                - set the column width to 6 
        * Freeze top row
        * Create and apply table formatting
        * Add explanatory texts after (right from) the generated table in case of "fake MultiIndex"

    Note: Empty A2 means the sheet has "fake MultiIndex" used as a compromize between excel and Backbone
    """
    wb = load_workbook(output_file)

    for ws in wb.worksheets:
        # Adjust each column's width
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 6  # add extra padding

        # Skip remaining processing if sheet has only 1 row
        if ws.max_row == 1:
            continue

        # If A2 is empty, the sheet has "fake MultiIndex" used as a compromize between excel and Backbone
        if ws["A2"].value is None:
            # Iterate cells in row 2 if cells are not empty
            for cell in ws[2]:
                if cell.value is not None:
                    # Rotate matching cell in row 1 if the length of the cell is more than 6 letters.
                    ws.cell(row=1, column=cell.col_idx).alignment = Alignment(textRotation=90)

                    # Centre align entire columns that have text in row 2 
                    col_index = cell.column
                    # Iterate over all cells in this column
                    for row_cells in ws.iter_rows(min_col=col_index, max_col=col_index):
                        for c in row_cells:
                            # Preserve any existing text rotation
                            current_rotation = c.alignment.textRotation if c.alignment else 0
                            c.alignment = Alignment(horizontal='center', textRotation=current_rotation)
                    col_letter = get_column_letter(cell.column)

                    # set the column width to 6 
                    ws.column_dimensions[col_letter].width = 6       

        # Special handling for 'p_gnu_io' sheet - replace zeros with empty strings in 'capacity' column
        if ws.title in ['p_gnu_io', 'p_unit', 'p_gn', 'p_gnBoundaryPropertiesForStates']:
            # Replace all zeros with empty strings in all columns of all sheets
            # Skip header row (row 1)
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value == 0 or cell.value == '0':
                        cell.value = ''

        # Freeze the top row
        ws.freeze_panes = "A2"

        # Derive table name from sheet name: remove any non-word characters and append _table.
        table_name = re.sub(r'\W+', '_', ws.title) + "_table"
        # Apply Excel table formatting
        last_col_letter = get_column_letter(ws.max_column)
        table_ref = f"A1:{last_col_letter}{ws.max_row}"
        table = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9",
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        table.headerRowCount = 1
        ws.add_table(table)


        # If A2 is empty, the sheet has "fake MultiIndex" used as a compromize between excel and Backbone
        if ws["A2"].value is None:
            # Add explanatory texts after (right from) the generated table
            n = ws.max_column + 2
            ws.cell(row=1, column=n, value='The first row labels are for excel Table headers.')
            ws.cell(row=2, column=n, value='The Second row labels are for GDXXRW converting excel to GDX.')


    # save the adjusted file
    wb.save(output_file)


def check_if_bb_excel_open(file_path):
    """
    Checks if the file at file_path is locked (e.g. currently open in Excel).
    If it is, raises an exception with an informative message.
    """
    if os.path.exists(file_path):
        try:
            with open(file_path, 'a'):
                pass
        except Exception as e:
            raise Exception(f"The Backbone input excel file '{file_path}' is currently open. Please close it and rerun the code.")
   
