"""Turning finished DataFrames into inputData.xlsx.

Everything here is about the *file*: the header row GDXXRW reads, the sheet
order, column widths, rotation, freeze panes, table styles, the index sheet.
None of it knows what a grid or a unit is, which is why it is not in
bb_excel_pipeline.py -- the builder decides what the model contains, this decides
what the workbook looks like.

The fake MultiIndex belongs here rather than with the frame helpers because it is
a way of *writing* a sheet, not of holding one. It is applied once per sheet on
the way out and nothing undoes it; there is no reader for it on this side.
"""

import os
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

import src.utils as utils


#: Sheets written with the fake MultiIndex, and the dimension columns that get a
#: blank rather than their name in its header row. Every other sheet is written as
#: it stands. These five are the ones src_files/indexSheet.xlsx declares with a
#: Cdim: p_gn Rdim=2/Cdim=1, p_gnn Rdim=3/Cdim=1, p_gnu_io Rdim=4/Cdim=1.
SHEET_DIMENSIONS = {
    'p_gn':                            ['grid', 'node'],
    'p_gnBoundaryPropertiesForStates': ['grid', 'node', 'param_gnBoundaryTypes'],
    'p_gnn':                           ['grid', 'from_node', 'to_node'],
    'p_gnu_io':                        ['grid', 'node', 'unit', 'input_output'],
    'p_unit':                          ['unit'],
}


def create_fake_MultiIndex(
    df: pd.DataFrame,
    dimensions: list[str]
    ) -> pd.DataFrame:
    """
    Creates a fake MultiIndex by:
    1. Taking an existing DataFrame with single-layer column names
    2. Creating a new first row with empty strings for dimensions and parameter names for parameter columns
    3. Shifting existing data down by one row

    Parameters:
    - df: pandas DataFrame with single-layered column names
    - dimensions: list of column names that are dimension columns

    Returns:
    - DataFrame with fake MultiIndex structure

    Dtypes stop meaning anything from here on, and here is the end
    --------------------------------------------------------------
    The inserted header row holds parameter *names* -- strings -- in every
    parameter column. So after this call every parameter column is ``object``
    by construction, no matter what it held before.

    This is the one place in the pipeline where ``object`` does **not** mean
    "no assumption has been made" (the all-NA rule in utils.py) and does not
    mean "something unparseable got in". It is simply what a sheet looks like
    once a text header row has been pushed into it.

    Nothing downstream has to cope with that, because there is no downstream:
    write_workbook calls this and writes the result. The builders work on plain
    frames whose dtypes still mean what utils.py says they mean.

    The numeric gate is deliberately **not** applied to the result. Every
    parameter column here legitimately mixes a name with numbers, which is
    exactly the shape gate_xlsx_frame exists to report; running it would
    flag every sheet in the workbook.
    """
    # Identify parameter columns (those not in dimensions)
    all_columns = list(df.columns)

    # Create a new DataFrame with the same columns
    df_output = pd.DataFrame(columns=all_columns)

    # Create the first row with empty strings for dimension columns
    # and parameter names for parameter columns
    first_row = []
    for col in all_columns:
        if col in dimensions:
            first_row.append("")
        else:
            first_row.append(col)

    # Add the first row to the DataFrame
    df_output.loc[0] = first_row

    # Reset the index of the input DataFrame and add it to the output
    df_reset = df.reset_index(drop=True)
    df_reset.index = df_reset.index + 1  # Shift indices to start from 1

    # Concatenate the first row with the original data
    df_output = pd.concat([df_output, df_reset], axis=0)
    df_output = utils.standardize_df_dtypes(df_output)

    return df_output


def write_workbook(output_file, sheets: dict) -> None:
    """Write one sheet per entry of `sheets`, in the order given.

    This is the only place the fake MultiIndex is applied. A builder holds an
    ordinary frame from its first row to its last, and a frame becomes a sheet
    GDXXRW can read exactly here -- which is why nothing needs to undo the
    transform, and why there is no drop_fake_MultiIndex to do it.

    A frame with no columns at all is written as an empty sheet rather than given
    a header row. That is a table the model does not have -- no transfer links, no
    units -- and a lone header row over nothing is not a truer statement of it.
    """
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, frame in sheets.items():
            dimensions = SHEET_DIMENSIONS.get(sheet_name)
            if dimensions is not None and len(frame.columns):
                frame = create_fake_MultiIndex(frame, dimensions)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)


def add_index_sheet(output_file, input_folder, logger) -> None:
    """
    Adds Index sheet to the excel
        * loads preconstructed 'indexSheet.xlsx'
        * picks rows where Symbol is in the sheet names
    """
    # An unset input folder needs no guard of its own: input_folder is a Path, so
    # comparing it with "" was never True, and the read below reports a missing
    # indexSheet.xlsx whatever the reason for it.
    index_path = os.path.join(input_folder, 'indexSheet.xlsx')

    # Read the index sheet file (assuming the first row contains headers)
    try:
        df_index = pd.read_excel(index_path, header=0)
    except Exception:
        logger.log_status(f"'{index_path}' not found, index sheet was not added to the BB input Excel.", level="warn")
        return

    # Load the output Excel workbook which already has multiple sheets
    wb = load_workbook(output_file)
    existing_sheet_names = wb.sheetnames

    # Filter rows: keep only rows where the 'Symbol' exists among the workbook's sheet names
    df_filtered = df_index[df_index['Symbol'].isin(existing_sheet_names)]

    # Create a new sheet named 'index'
    new_sheet = wb.create_sheet(title='index')

    # Write header row (row 1)
    for col_num, header in enumerate(df_index.columns, start=1):
        new_sheet.cell(row=1, column=col_num, value=header)

    # Write the filtered data starting from row 2
    for row_num, row in enumerate(df_filtered.itertuples(index=False, name=None), start=2):
        for col_num, value in enumerate(row, start=1):
            new_sheet.cell(row=row_num, column=col_num, value=value)

    # Move the 'index' sheet to the first position in the workbook
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(new_sheet)))

    # Save the updated workbook back to the output file
    wb.save(output_file)


def adjust_excel(output_file) -> None:
    """
    For each sheet in the Excel file
        * Adjust each column's width.
        * Skip remaining processing if sheet has only 1 row.
        * If A2 is empty, iterate non-empty cells in row 2:
            - Rotate matching cell in row 1 if the length of the cell is more than 6 letters.
            - Centre align columns
            - set the column width to 6
        * Freeze top row
        * Create and apply table formatting
        * Add explanatory texts after (right from) the generated table in case of "fake MultiIndex"

    Note: Empty A2 means the sheet has "fake MultiIndex" used as a compromize between excel and Backbone
    """
    wb = load_workbook(output_file)

    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column

        # Adjust each column's width (based on longest value in column)
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0

            for row_idx in range(1, max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    length = len(str(value))
                    if length > max_length:
                        max_length = length

            ws.column_dimensions[col_letter].width = max_length + 6  # padding

        # Skip remaining processing if sheet has only 1 row
        if ws.max_row == 1:
            continue

        # If A2 is empty, the sheet has "fake MultiIndex" used as a compromize between excel and GDXXRW
        has_fake_multiindex = ws["A2"].value is None
        if has_fake_multiindex:
            # Pre-create alignments to avoid recreating them in loops
            center_align = Alignment(horizontal="center")
            rotated_header_align = Alignment(horizontal="center", textRotation=90)

            # Iterate cells in row 2 if cells are not empty
            for cell in ws[2]:
                if cell.value is None:
                    continue

                col_idx = cell.col_idx
                col_letter = get_column_letter(col_idx)

                # Rotate matching cell in row 1 if the length of the cell is more than 6 letters.
                header_cell = ws.cell(row=1, column=col_idx)
                header_text = str(header_cell.value) if header_cell.value is not None else ""
                if len(header_text) > 6:
                    header_cell.alignment = rotated_header_align

                # Centre align column values from row 2 downwards
                for row_idx in range(2, max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).alignment = center_align

                # Set the column width to 6 for these rotated / "special" columns
                ws.column_dimensions[col_letter].width = 6

        # Freeze the top row
        ws.freeze_panes = "A2"

        # Create and apply table formatting
        # Derive table name from sheet name: remove any non-word characters and append _table.
        table_name = re.sub(r'\W+', '_', ws.title) + "_table"
        # Apply Excel table formatting
        last_col_letter = get_column_letter(ws.max_column)
        table_ref = f"A1:{last_col_letter}{ws.max_row}"
        table = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9",
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        table.headerRowCount = 1
        ws.add_table(table)

        # If fake MultiIndex, add explanatory texts to the right of the table
        if has_fake_multiindex:
            n = ws.max_column + 2
            ws.cell(row=1, column=n, value='The first row labels are for excel Table headers.')
            ws.cell(row=2, column=n, value='The Second row labels are for GDXXRW converting excel to GDX.')

    # save the adjusted file
    wb.save(output_file)
