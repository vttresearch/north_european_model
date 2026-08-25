"""The ``.wb.txt`` fixture format: source workbooks stored as text.

Why text
--------
Test workbooks used to be a hand-edited binary file. It could only be changed by
opening Excel, a reviewer could not see what a change did, and it silently rotted
until every one of its sheets truncated to zero rows. Storing fixtures as text
makes them diffable, reviewable, and impossible to break without the diff showing
it.

Why a sectioned file rather than one CSV per sheet
--------------------------------------------------
The two highest-value behaviours of the reader are *about blank lines and hash
lines*: ``read_input_excels`` truncates a sheet at the first fully-empty row
(source_data_loader.py:102-107) and ``normalize_dataframe`` drops rows whose
cells start with ``#`` (:201-208). CSV cannot express either without an
out-of-band convention, at which point it is no longer just CSV. Here ``//`` is
the fixture's own comment marker, so ``#`` passes straight through to Excel and
stays testable, and ``<<blank>>`` is the only way to emit a blank row -- which no
editor or formatter can helpfully "clean up".

One way only
------------
The ``.wb.txt`` is the source of truth; the ``.xlsx`` is a ``tmp_path`` build
artifact. Editing a generated workbook must lose the edit, otherwise fixtures
drift back out of the repo, which is exactly how the old one died.
``dump_xlsx_to_text`` exists for bootstrapping a fixture from a real workbook,
not for round-tripping.

Grammar
-------
=========================  ====================================================
``// text``                fixture comment; never reaches Excel
blank line                 formatting only; ignored
``[sheetname]``            starts a sheet; the name is used verbatim, so
                           ``[unitdata_FI]`` exercises prefix matching
first line after ``[...]`` header row
``|``                      field separator; every field is stripped
empty field / trailing ``|``  a genuinely blank cell (``None``)
``<<blank>>``              one all-blank row -- the only way to write one
leading ``'``              force a text cell (``'2030``), Excel's own convention
fewer fields than header   trailing cells blank
more fields than header    parse error naming the file and line
=========================  ====================================================

Values are typed the way a human typing into Excel would produce them: blank
becomes ``None``, a ``'`` prefix forces text, otherwise int, then float, then
str. There is deliberately **no dtype-declaration syntax** -- ``standardize_df_dtypes``
is itself under test, so a fixture stating ``capacity: Float64`` would pre-decide
the thing the pipeline decides, and would be the banned per-column dtype map
under another name.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook, load_workbook

#: Marker for a deliberately blank spreadsheet row.
BLANK_ROW = "<<blank>>"

#: Fixture comment prefix. Distinct from '##', which is meaningful *to the
#: pipeline* -- it is how an author marks a row or column as not being input --
#: and must therefore survive into the workbook rather than being eaten here.
COMMENT_PREFIX = "//"

WORKBOOK_SUFFIX = ".wb.txt"

Rows = list[list[Any]]
Sheets = dict[str, Rows]

_FIXTURE_DIR = Path(__file__).resolve().parent / "workbooks"


class WorkbookTextError(ValueError):
    """A ``.wb.txt`` fixture could not be parsed."""


def _coerce(field: str) -> Any:
    """Type one field the way Excel would if a human typed it."""
    if field == "":
        return None
    if field.startswith("'"):
        return field[1:]
    try:
        return int(field)
    except ValueError:
        pass
    try:
        return float(field)
    except ValueError:
        pass
    return field


def parse_workbook_text(text: str, *, source: str = "<string>") -> Sheets:
    """Parse `.wb.txt` content into ``{sheet_name: rows}``; ``rows[0]`` is the header."""
    sheets: Sheets = {}
    current: str | None = None

    for lineno, raw in enumerate(text.splitlines(), start=1):
        line = raw.strip()
        if not line or line.startswith(COMMENT_PREFIX):
            continue

        if line.startswith("[") and line.endswith("]"):
            current = line[1:-1].strip()
            if not current:
                raise WorkbookTextError(f"{source}:{lineno}: empty sheet name")
            if current in sheets:
                raise WorkbookTextError(
                    f"{source}:{lineno}: sheet {current!r} declared twice"
                )
            sheets[current] = []
            continue

        if current is None:
            raise WorkbookTextError(
                f"{source}:{lineno}: data before any [sheet] header -- {line!r}"
            )

        rows = sheets[current]
        if line == BLANK_ROW:
            if not rows:
                raise WorkbookTextError(
                    f"{source}:{lineno}: {BLANK_ROW} cannot be the header row"
                )
            rows.append([None] * len(rows[0]))
            continue

        fields = [f.strip() for f in line.split("|")]
        if rows and len(fields) > len(rows[0]):
            raise WorkbookTextError(
                f"{source}:{lineno}: {len(fields)} fields but the header has "
                f"{len(rows[0])} -- {line!r}"
            )
        values = [_coerce(f) for f in fields]
        if rows:
            values += [None] * (len(rows[0]) - len(values))
        rows.append(values)

    return sheets


def sheet_names(text: str) -> list[str]:
    """Declared sheet names, without building anything.

    ``run_route`` uses this to derive the config's ``*_files`` keys, so a new
    data category is picked up as soon as a fixture declares a sheet for it.
    """
    return [
        line.strip()[1:-1].strip()
        for line in text.splitlines()
        if line.strip().startswith("[") and line.strip().endswith("]")
    ]


def write_workbook(sheets: Sheets, path: Path) -> Path:
    """Write ``{sheet: rows}`` to an .xlsx.

    openpyxl directly, not ``pandas.ExcelWriter``: pandas would re-infer dtypes
    and reintroduce exactly the typing opinions this format refuses to hold.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)  # drop the default empty sheet
    for name, rows in sheets.items():
        worksheet = workbook.create_sheet(title=name)
        for row in rows:
            worksheet.append(row)
    workbook.save(path)
    return path


def write_workbook_text(text: str, path: Path, *, source: str | None = None) -> Path:
    """Parse `.wb.txt` content and write it as an .xlsx."""
    return write_workbook(parse_workbook_text(text, source=source or str(path)), path)


def load_workbook_fixture(name: str) -> str:
    """Read ``tests/_common/workbooks/<name>.wb.txt``. The suffix is optional."""
    filename = name if name.endswith(WORKBOOK_SUFFIX) else f"{name}{WORKBOOK_SUFFIX}"
    path = _FIXTURE_DIR / filename
    if not path.is_file():
        available = sorted(p.name for p in _FIXTURE_DIR.glob(f"*{WORKBOOK_SUFFIX}"))
        raise FileNotFoundError(f"no fixture {filename!r}; available: {available}")
    return path.read_text(encoding="utf-8")


def _row_matches(header: Sequence[str], fields: Sequence[str], where: Mapping[str, Any]) -> bool:
    for key, wanted in where.items():
        if key not in header:
            return False
        idx = header.index(key)
        if idx >= len(fields):
            return False
        if fields[idx].strip().lstrip("'") != str(wanted):
            return False
    return True


def workbook_text_with(
    text: str,
    *,
    sheet: str,
    header: str,
    value: Any,
    where: Mapping[str, Any],
) -> str:
    """Return `text` with exactly ONE cell changed. The delta primitive.

    `where` selects the row by ``{header_name: cell_value}``. Selecting zero rows
    or more than one is an error, not a silent no-op: an edit that quietly
    changes nothing is the single way a delta test can pass for the wrong reason,
    so it has to be impossible.

    Comments and the alignment of untouched lines are preserved, so the diff
    shows the one cell that moved.
    """
    lines = text.splitlines()
    out = list(lines)

    current: str | None = None
    # Held across later sections: a sheet that appears before others must still
    # be recognised as found once the loop has moved past it.
    header_fields: list[str] | None = None
    seen_header = False
    matches: list[int] = []

    for index, raw in enumerate(lines):
        line = raw.strip()
        if not line or line.startswith(COMMENT_PREFIX):
            continue
        if line.startswith("[") and line.endswith("]"):
            current = line[1:-1].strip()
            continue
        if current != sheet:
            continue
        if not seen_header:
            header_fields = [f.strip() for f in line.split("|")]
            seen_header = True
            if header not in header_fields:
                raise WorkbookTextError(
                    f"sheet {sheet!r} has no column {header!r}; columns are "
                    f"{header_fields}"
                )
            continue
        if line == BLANK_ROW:
            continue
        fields = [f.strip() for f in line.split("|")]
        if _row_matches(header_fields, fields, where):
            matches.append(index)

    if not seen_header:
        raise WorkbookTextError(f"no sheet named {sheet!r} in this fixture")
    if not matches:
        raise WorkbookTextError(
            f"no row in sheet {sheet!r} matches {dict(where)!r}; a variant that "
            f"changes nothing would make a delta test pass for the wrong reason"
        )
    if len(matches) > 1:
        raise WorkbookTextError(
            f"{len(matches)} rows in sheet {sheet!r} match {dict(where)!r}; the "
            f"selection must identify exactly one row"
        )

    index = matches[0]
    column = header_fields.index(header)
    segments = out[index].split("|")
    replacement = "" if value is None else str(value)
    # Keep the original field width so the table stays aligned in the diff.
    original = segments[column]
    leading = len(original) - len(original.lstrip())
    segments[column] = (" " * leading) + replacement.ljust(
        max(0, len(original) - leading)
    )
    out[index] = "|".join(segments)
    return "\n".join(out) + ("\n" if text.endswith("\n") else "")


def dump_xlsx_to_text(path: Path) -> str:
    """Render an existing .xlsx as `.wb.txt`. Bootstrapping only, not a round-trip."""
    workbook = load_workbook(Path(path), data_only=True)
    blocks: list[str] = []
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        widths: dict[int, int] = {}
        rendered: list[list[str]] = []
        for row in rows:
            cells = ["" if v is None else str(v) for v in row]
            rendered.append(cells)
            for i, cell in enumerate(cells):
                widths[i] = max(widths.get(i, 0), len(cell))
        lines = [f"[{worksheet.title}]"]
        for cells in rendered:
            if not any(cells):
                lines.append(BLANK_ROW)
                continue
            lines.append(
                " | ".join(cell.ljust(widths.get(i, 0)) for i, cell in enumerate(cells)).rstrip()
            )
        blocks.append("\n".join(lines))
    return "\n\n".join(blocks) + "\n"
