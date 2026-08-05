"""Tests for the ``.wb.txt`` fixture format.

The format's whole justification is that it can express things CSV cannot --
``#`` comment rows and blank truncation rows, both of which the reader treats as
meaningful. Those cases get the most attention here.

``workbook_text_with`` gets negative controls: it is the delta primitive, and a
variant that silently changes nothing would make every delta test pass for the
wrong reason.
"""

import pandas as pd
import pytest
from openpyxl import load_workbook

from tests._common.workbook_text import (
    BLANK_ROW,
    WorkbookTextError,
    dump_xlsx_to_text,
    load_workbook_fixture,
    parse_workbook_text,
    sheet_names,
    workbook_text_with,
    write_workbook_text,
)

SIMPLE = """\
// a fixture comment
[unitdata]
Country | Generator_ID | capacity
FI      | coal         | 100
SE      | wind         | 200
"""


class TestParsing:
    def test_reads_sheets_and_rows(self):
        sheets = parse_workbook_text(SIMPLE)
        assert list(sheets) == ["unitdata"]
        assert sheets["unitdata"][0] == ["Country", "Generator_ID", "capacity"]
        assert sheets["unitdata"][1] == ["FI", "coal", 100]

    def test_fixture_comments_never_reach_the_workbook(self):
        sheets = parse_workbook_text(SIMPLE)
        assert not any("fixture comment" in str(cell) for row in sheets["unitdata"] for cell in row)

    def test_a_hash_row_survives_because_the_pipeline_gives_it_meaning(self):
        """``#`` is the pipeline's comment marker, not the fixture's.

        normalize_dataframe drops rows whose cells start with '#'
        (source_data_loader.py:201-208). If the fixture format ate them, that
        behaviour would be untestable -- which is the main reason '//' is the
        fixture comment.
        """
        sheets = parse_workbook_text(
            "[unitdata]\nCountry | capacity\n#note   | 1\nFI      | 2\n"
        )
        assert sheets["unitdata"][1] == ["#note", 1]

    def test_blank_marker_emits_an_all_blank_row(self):
        # The reader truncates a sheet at the first fully-empty row, so this has
        # to be expressible -- and no formatter can strip a <<blank>> token.
        sheets = parse_workbook_text(
            f"[unitdata]\nCountry | capacity\nFI | 1\n{BLANK_ROW}\nSE | 2\n"
        )
        assert sheets["unitdata"][2] == [None, None]
        assert sheets["unitdata"][3] == ["SE", 2]

    def test_ordinary_blank_lines_are_only_formatting(self):
        sheets = parse_workbook_text("[unitdata]\nCountry\n\nFI\n\n\nSE\n")
        assert sheets["unitdata"] == [["Country"], ["FI"], ["SE"]]

    @pytest.mark.parametrize(
        "field, expected",
        [
            ("100", 100),
            ("1.5", 1.5),
            ("-2", -2),
            ("1e3", 1000.0),
            ("FI", "FI"),
            ("", None),
            ("'2030", "2030"),
            ("'0", "0"),
        ],
        ids=["int", "float", "negative", "exponent", "text", "blank", "forced", "forced-zero"],
    )
    def test_field_typing(self, field, expected):
        # Two columns, because in a one-column sheet an empty field and a blank
        # line are the same characters -- see the test below.
        sheets = parse_workbook_text(f"[s]\ncol | other\n{field} | x\n")
        assert sheets["s"][1][0] == expected

    def test_a_lone_blank_cell_needs_the_blank_marker(self):
        """Documented limitation, not a bug.

        In a single-column sheet an empty field is literally an empty line, and
        empty lines are formatting. ``<<blank>>`` is the unambiguous way to say
        "a blank row", which is why it exists.
        """
        assert parse_workbook_text("[s]\ncol\n\n") == {"s": [["col"]]}
        assert parse_workbook_text(f"[s]\ncol\n{BLANK_ROW}\n") == {"s": [["col"], [None]]}

    def test_a_trailing_separator_makes_a_blank_cell(self):
        sheets = parse_workbook_text("[s]\na | b\nFI |\n")
        assert sheets["s"][1] == ["FI", None]

    def test_short_rows_are_padded(self):
        sheets = parse_workbook_text("[s]\na | b | c\nFI\n")
        assert sheets["s"][1] == ["FI", None, None]

    def test_sheet_names_are_verbatim_so_prefix_matching_is_testable(self):
        # Sheets are selected by case-insensitive prefix, so a fixture must be
        # able to declare unitdata_FI and unitdata_SE separately.
        assert sheet_names("[unitdata_FI]\na\n[unitdata_SE]\na\n") == [
            "unitdata_FI",
            "unitdata_SE",
        ]


class TestParseErrors:
    def test_more_fields_than_the_header(self):
        with pytest.raises(WorkbookTextError, match="3 fields but the header has 2"):
            parse_workbook_text("[s]\na | b\n1 | 2 | 3\n")

    def test_the_error_names_the_file_and_line(self):
        with pytest.raises(WorkbookTextError, match=r"chp\.wb\.txt:3"):
            parse_workbook_text("[s]\na | b\n1 | 2 | 3\n", source="chp.wb.txt")

    def test_data_before_any_sheet_header(self):
        with pytest.raises(WorkbookTextError, match="before any \\[sheet\\]"):
            parse_workbook_text("Country | capacity\nFI | 1\n")

    def test_a_duplicate_sheet_name(self):
        with pytest.raises(WorkbookTextError, match="declared twice"):
            parse_workbook_text("[s]\na\n[s]\na\n")

    def test_an_empty_sheet_name(self):
        with pytest.raises(WorkbookTextError, match="empty sheet name"):
            parse_workbook_text("[]\na\n")

    def test_blank_marker_as_the_header(self):
        with pytest.raises(WorkbookTextError, match="cannot be the header"):
            parse_workbook_text(f"[s]\n{BLANK_ROW}\n")


class TestWriting:
    def test_produces_a_workbook_the_pipeline_reader_can_open(self, tmp_path):
        path = write_workbook_text(SIMPLE, tmp_path / "data.xlsx")
        frame = pd.read_excel(path, sheet_name="unitdata", header=0)
        assert list(frame.columns) == ["Country", "Generator_ID", "capacity"]
        assert frame["capacity"].tolist() == [100, 200]

    def test_blank_rows_are_genuinely_empty_cells(self, tmp_path):
        # Not the empty string: read_input_excels truncates on a fully-empty row,
        # and "" would not count.
        path = write_workbook_text(
            f"[s]\na | b\nFI | 1\n{BLANK_ROW}\nSE | 2\n", tmp_path / "d.xlsx"
        )
        rows = list(load_workbook(path)["s"].iter_rows(values_only=True))
        assert rows[2] == (None, None)

    def test_a_blank_row_truncates_the_sheet_for_the_real_reader(self, tmp_path):
        """The behaviour the format exists to make testable, end to end."""
        from src.source_data.source_data_loader import read_input_excels
        from tests._common.fixtures import FakeLogger

        folder = tmp_path / "data_files"
        write_workbook_text(
            f"[unitdata]\nCountry | capacity\nFI | 1\n{BLANK_ROW}\nSE | 2\n",
            folder / "d.xlsx",
        )
        frames = read_input_excels(folder, ["d.xlsx"], "unitdata", FakeLogger())

        assert len(frames) == 1
        assert frames[0]["Country"].tolist() == ["FI"]  # SE is below the blank row

    def test_forced_text_stays_text(self, tmp_path):
        path = write_workbook_text("[s]\nyear\n'2030\n", tmp_path / "d.xlsx")
        assert load_workbook(path)["s"]["A2"].value == "2030"

    def test_sheet_names_survive(self, tmp_path):
        path = write_workbook_text("[unitdata_FI]\na\n1\n[nodedata]\nb\n2\n", tmp_path / "d.xlsx")
        assert load_workbook(path).sheetnames == ["unitdata_FI", "nodedata"]


class TestWorkbookTextWith:
    def test_changes_exactly_one_cell(self):
        out = workbook_text_with(
            SIMPLE, sheet="unitdata", header="capacity", value=750,
            where={"Country": "FI"},
        )
        sheets = parse_workbook_text(out)
        assert sheets["unitdata"][1] == ["FI", "coal", 750]
        assert sheets["unitdata"][2] == ["SE", "wind", 200]   # untouched

    def test_preserves_comments(self):
        out = workbook_text_with(
            SIMPLE, sheet="unitdata", header="capacity", value=750,
            where={"Country": "FI"},
        )
        assert "// a fixture comment" in out

    def test_can_blank_a_cell(self):
        out = workbook_text_with(
            SIMPLE, sheet="unitdata", header="capacity", value=None,
            where={"Country": "FI"},
        )
        assert parse_workbook_text(out)["unitdata"][1] == ["FI", "coal", None]

    def test_matches_on_several_columns(self):
        text = "[s]\na | b | v\nFI | x | 1\nFI | y | 2\n"
        out = workbook_text_with(text, sheet="s", header="v", value=9,
                                 where={"a": "FI", "b": "y"})
        assert parse_workbook_text(out)["s"][2] == ["FI", "y", 9]

    def test_raises_when_nothing_matches(self):
        # THE negative control. A silently-no-op edit is the one way a delta test
        # can pass for the wrong reason.
        with pytest.raises(WorkbookTextError, match="no row in sheet"):
            workbook_text_with(SIMPLE, sheet="unitdata", header="capacity",
                               value=1, where={"Country": "DK"})

    def test_raises_when_several_rows_match(self):
        text = "[s]\na | v\nFI | 1\nFI | 2\n"
        with pytest.raises(WorkbookTextError, match="2 rows in sheet"):
            workbook_text_with(text, sheet="s", header="v", value=9, where={"a": "FI"})

    def test_finds_a_sheet_that_is_not_the_last_one(self):
        # Regression: the header tracker was reset on entering each new section,
        # so a target sheet followed by any other sheet reported "no sheet named".
        text = "[unitdata]\na | v\nFI | 1\n\n[nodedata]\nb\nx\n"
        out = workbook_text_with(text, sheet="unitdata", header="v", value=9,
                                 where={"a": "FI"})
        assert parse_workbook_text(out)["unitdata"][1] == ["FI", 9]

    def test_does_not_match_rows_in_a_different_sheet(self):
        # The other half: identical column names in two sheets must not collide.
        text = "[unitdata]\na | v\nFI | 1\n\n[nodedata]\na | v\nFI | 2\n"
        out = workbook_text_with(text, sheet="nodedata", header="v", value=9,
                                 where={"a": "FI"})
        sheets = parse_workbook_text(out)
        assert sheets["unitdata"][1] == ["FI", 1]   # untouched
        assert sheets["nodedata"][1] == ["FI", 9]

    def test_raises_on_an_unknown_sheet(self):
        with pytest.raises(WorkbookTextError, match="no sheet named"):
            workbook_text_with(SIMPLE, sheet="nodedata", header="x", value=1, where={})

    def test_raises_on_an_unknown_column(self):
        with pytest.raises(WorkbookTextError, match="no column"):
            workbook_text_with(SIMPLE, sheet="unitdata", header="nonexistent",
                               value=1, where={"Country": "FI"})

    def test_the_result_still_parses(self):
        out = workbook_text_with(SIMPLE, sheet="unitdata", header="Generator_ID",
                                 value="a much longer generator name",
                                 where={"Country": "FI"})
        assert parse_workbook_text(out)["unitdata"][1][1] == "a much longer generator name"


class TestDumpXlsxToText:
    def test_round_trips_values_for_bootstrapping(self, tmp_path):
        path = write_workbook_text(SIMPLE, tmp_path / "d.xlsx")
        assert parse_workbook_text(dump_xlsx_to_text(path)) == parse_workbook_text(SIMPLE)

    def test_renders_blank_rows_as_the_marker(self, tmp_path):
        path = write_workbook_text(
            f"[s]\na | b\nFI | 1\n{BLANK_ROW}\nSE | 2\n", tmp_path / "d.xlsx"
        )
        assert BLANK_ROW in dump_xlsx_to_text(path)


class TestFixtureLoading:
    def test_an_unknown_fixture_lists_what_is_available(self):
        with pytest.raises(FileNotFoundError, match="available:"):
            load_workbook_fixture("does_not_exist")
