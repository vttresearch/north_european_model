import shutil
import sys
import os
import build_input_data
from pathlib import Path
from src.config_reader import load_config
from openpyxl import load_workbook

if __name__ == '__main__':
    input_folder = sys.argv[1]
    config_files = sys.argv[2:]
    for config_file in config_files:
        #get the folder name
        config = load_config(Path(config_file))
        output_folder_prefix = config.get('output_folder_prefix')
        scenarios = config.get('scenarios')
        scenario_years = config.get('scenario_years')
        scenario_alternatives = config.get('scenario_alternatives')
        output_folders = []
        for scenario in scenarios:
            for scenario_year in scenario_years:
                if scenario_alternatives and scenario_alternatives != ['']:
                    for scenario_alternative in scenario_alternatives:
                        output_folders.append(os.path.join(f"{output_folder_prefix}_{scenario}_{str(scenario_year)}_{scenario_alternative}"))
                else:
                    output_folders.append(os.path.join(f"{output_folder_prefix}_{scenario}_{str(scenario_year)}"))

        #call the actual script
        build_input_data.main(input_folder, Path(config_file))
        
        for output_folder in output_folders:
            print(output_folder)
            #copy the inputdata excel to another folder for workflow use
            os.makedirs(f'./toolbox_workflow/input', exist_ok=True)
            output_name = f'./toolbox_workflow/input/{output_folder}_inputData.xlsx'.replace(" ", "")
            shutil.copy(f'./{output_folder}/inputData.xlsx', output_name)
            
            #add folder information
            wb = load_workbook(output_name)          
            table_name = "scenario_folder"
            wb.create_sheet(table_name)
            for i, name in enumerate(wb.sheetnames):
                if name == "index":
                    last_row = len(wb.worksheets[i]['A'])
                    wb.worksheets[i].cell(row=last_row + 1, column=1).value = "Set"
                    wb.worksheets[i].cell(row=last_row + 1, column=2).value = table_name
                    wb.worksheets[i].cell(row=last_row + 1, column=3).value = f'{table_name}!A2'
                    wb.worksheets[i].cell(row=last_row + 1, column=4).value = 1
                if name == table_name:
                    wb.worksheets[i].cell(row=1, column=1).value = table_name
                    wb.worksheets[i].cell(row=2, column=1).value = output_folder
            wb.save(output_name)